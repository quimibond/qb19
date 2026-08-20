# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Mapa de celdas fijas del formato actual de "Ficha Técnica de Proceso Tejido
# Circular" de Quimibond. Si el formato de origen cambia, solo hay que
# actualizar este diccionario — el resto del wizard no requiere tocarse.
CELL_MAP = {
    'articulo': 'G9',
    'maquina_tejido': 'AD9',
    'marca_maquina': 'H28',
    'galga': 'AF28',
    'diametro': 'AF29',
    'no_agujas': 'AF30',
    'no_alimentadores': 'AT28',
    'velocidad_raw': 'AR9',       # ej. " 24.90 rpm/min" -> se parsea el número
    'vueltas_por_rollo': 'AP11',
    'longitud_malla': 'K36',
    'longitud_malla_tol': 'W36',
    'consumo_cm_vta': 'K37',
    'consumo_cm_vta_tol': 'W37',
    'polea_alimentacion': 'K38',
    'tension': 'K39',
    'punto_cilindro': 'K40',
    'punto_plato': 'K41',
    'altura_plato': 'K42',
    'ancho_bastidor': 'K43',
    'ancho_bastidor_tol': 'W43',
    'estiraje': 'K44',
    'ancho_rollo': 'K45',
    'ancho_rollo_tol': 'W45',
    'peso_promedio_rollo': 'K46',
    'peso_promedio_rollo_tol': 'W46',
    'peso_acondicionado': 'U60',
    'ancho_acondicionado': 'U61',
    'espesor_acondicionado': 'U62',
    'columnas': 'U63',
    'mallas': 'U64',
    'elongacion_carga_largo': 'U65',
    'elongacion_carga_ancho': 'U66',
    'jefe_manufactura': 'F76',
    'auxiliar_procesos': 'U76',
}

# Filas de la tabla de disposición de hilo (columnas fijas E/I/O/U/Z/AD)
HILO_ROWS = [15, 16, 17, 18]
HILO_COLS = {
    'numero': 'E', 'tipo_hilo': 'I', 'titulo_hilo': 'O',
    'torsion': 'U', 'porcentaje': 'Z', 'lote': 'AD',
}


def _clean_num(value):
    if value in (None, '', 'N/A'):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        digits = ''.join(ch for ch in str(value) if (ch.isdigit() or ch in '.-'))
        return float(digits) if digits else 0.0
    except ValueError:
        return 0.0


class FichaTecnicaImportWizard(models.TransientModel):
    _name = 'ficha.tecnica.import.wizard'
    _description = 'Importar Ficha Técnica de Tela desde Excel'

    file_data = fields.Binary(string='Archivo Excel (.xlsx)', required=True)
    file_name = fields.Char(string='Nombre de archivo')
    product_proceso_id = fields.Many2one(
        'product.product', string='Producto — Tela en Proceso (kg)',
        help='Opcional. Si se indica, se vincula a la ficha importada.')
    product_acabado_id = fields.Many2many(
        'product.product', string='Productos — Tela Acabada (m)',
        help='Opcional. Si se indican, se vinculan a la ficha importada '
             '(ej. distintos colores de la misma construcción).')
    update_if_exists = fields.Boolean(
        string='Actualizar si ya existe (mismo artículo + revisión)', default=True)

    def action_import(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(
                'La librería openpyxl no está disponible en este servidor. '
                'Solicite a su administrador que la instale para poder '
                'importar fichas técnicas desde Excel.')

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(self.file_data)), data_only=True)
        except Exception as exc:
            raise UserError('No se pudo leer el archivo. Verifique que sea un '
                             '.xlsx válido con el formato de ficha técnica de '
                             'Quimibond. Detalle: %s' % exc)

        ws = wb.active
        values = {}
        for field_name, coord in CELL_MAP.items():
            cell_value = ws[coord].value
            if field_name == 'velocidad_raw':
                values['velocidad'] = _clean_num(cell_value)
            elif field_name in ('galga', 'longitud_malla', 'longitud_malla_tol',
                                 'consumo_cm_vta', 'consumo_cm_vta_tol',
                                 'ancho_bastidor', 'ancho_bastidor_tol',
                                 'ancho_rollo', 'ancho_rollo_tol',
                                 'peso_promedio_rollo', 'peso_promedio_rollo_tol',
                                 'peso_acondicionado', 'ancho_acondicionado',
                                 'espesor_acondicionado', 'elongacion_carga_largo',
                                 'elongacion_carga_ancho'):
                values[field_name] = _clean_num(cell_value)
            elif field_name in ('no_agujas', 'no_alimentadores', 'vueltas_por_rollo',
                                 'columnas', 'mallas'):
                values[field_name] = int(_clean_num(cell_value))
            else:
                values[field_name] = cell_value

        if not values.get('articulo'):
            raise UserError(
                'No se encontró el código de artículo en la celda esperada (%s). '
                'Verifique que el archivo respete el formato estándar de ficha '
                'técnica de tejido.' % CELL_MAP['articulo'])

        # Líneas de disposición de hilo
        hilo_lines = []
        for row in HILO_ROWS:
            tipo = ws['%s%s' % (HILO_COLS['tipo_hilo'], row)].value
            if not tipo or tipo == 'N/A':
                continue
            hilo_lines.append((0, 0, {
                'numero': ws['%s%s' % (HILO_COLS['numero'], row)].value,
                'tipo_hilo': tipo,
                'titulo_hilo': ws['%s%s' % (HILO_COLS['titulo_hilo'], row)].value,
                'torsion': ws['%s%s' % (HILO_COLS['torsion'], row)].value,
                'porcentaje': _clean_num(ws['%s%s' % (HILO_COLS['porcentaje'], row)].value),
                'lote': ws['%s%s' % (HILO_COLS['lote'], row)].value,
            }))
        values['hilo_line_ids'] = hilo_lines

        if self.product_proceso_id:
            values['product_proceso_id'] = self.product_proceso_id.id
        if self.product_acabado_id:
            values['product_acabado_id'] = [(6, 0, self.product_acabado_id.ids)]

        Ficha = self.env['ficha.tecnica.tela']
        existing = Ficha.search([
            ('articulo', '=', values['articulo']),
            ('revision', '=', values.get('revision', '0') or '0'),
        ], limit=1)

        if existing:
            if not self.update_if_exists:
                raise UserError(
                    'Ya existe una ficha técnica para el artículo %s (revisión %s). '
                    'Marque "Actualizar si ya existe" para sobrescribirla.'
                    % (values['articulo'], values.get('revision', '0')))
            values['hilo_line_ids'] = [(5, 0, 0)] + hilo_lines
            existing.write(values)
            ficha = existing
        else:
            ficha = Ficha.create(values)

        return {
            'type': 'ir.actions.act_window',
            'name': 'Ficha Técnica de Tela',
            'res_model': 'ficha.tecnica.tela',
            'view_mode': 'form',
            'res_id': ficha.id,
        }
