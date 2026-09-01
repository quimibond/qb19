# -*- coding: utf-8 -*-
"""Presupuesto maestro de ventas — Paso 1: modelo, unidades, divisas y real.

Datos propios (no demo). Periodo 2040 para aislar de la facturación de demo.
"""
import base64
import io
from datetime import date

from odoo import fields
from odoo.tests import TransactionCase, tagged
from odoo.exceptions import UserError, ValidationError


def _seed_budget_pricelist(env, price):
    """Configura la lista presupuestal (P5) con una regla GLOBAL fija: valúa las
    líneas del presupuesto SIN cliente (antes se tomaba una lista arbitraria)."""
    pl = env['product.pricelist'].create({'name': 'Presupuestal test'})
    env['product.pricelist.item'].create({
        'pricelist_id': pl.id, 'applied_on': '3_global',
        'compute_price': 'fixed', 'fixed_price': price})
    env['ir.config_parameter'].sudo().set_param(
        'quimibond_sgi.budget_pricelist_id', pl.id)
    return pl


@tagged('post_install', '-at_install')
class TestSalesBudget(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.company = cls.env.company
        cls.team = cls.env['crm.team'].create({'name': 'Mercado Industrial KPI'})
        # Por xmlid (el nombre depende del idioma): m/cm = Longitud, kg = Peso.
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.uom_cm = cls.env.ref('uom.product_uom_cm')
        cls.uom_kg = cls.env.ref('uom.product_uom_kgm')
        cls.product = cls.env['product.product'].create({
            'name': 'Tela industrial PPV', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        cls.partner = cls.env['res.partner'].create({'name': 'Cliente PPV'})
        # Lista presupuestal (P5) para valuar las líneas globales a 20.
        _seed_budget_pricelist(cls.env, 20.0)

    def _budget(self, **vals):
        base = {'year': 2040, 'team_id': self.team.id}
        base.update(vals)
        return self.Budget.create(base)

    def _line(self, budget, when, uom=None, qty=0.0, amount=0.0, product=None):
        return self.Line.create({
            'budget_id': budget.id, 'product_id': (product or self.product).id,
            'date': when, 'uom_id': (uom or self.uom_m).id,
            'qty_budget': qty, 'amount_budget': amount})

    def _invoice(self, when, qty, uom, price, refund=False, currency=None):
        move = self.env['account.move'].create({
            'move_type': 'out_refund' if refund else 'out_invoice',
            'partner_id': self.partner.id, 'invoice_date': when,
            'team_id': self.team.id,
            'currency_id': (currency or self.company.currency_id).id,
            'invoice_line_ids': [(0, 0, {
                'product_id': self.product.id, 'quantity': qty,
                'product_uom_id': uom.id, 'price_unit': price,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    # ------------------------------------------------------------------
    def test_01_name_and_currency(self):
        budget = self._budget()
        self.assertEqual(budget.name, "Presupuesto Mercado Industrial KPI 2040 Rev.1")
        self.assertTrue(budget.folio.startswith('PPV-'))
        self.assertEqual(budget.currency_id, self.company.currency_id)

    def test_02_unique_active_per_year_team(self):
        self._budget()
        with self.assertRaises(ValidationError):
            self._budget()  # segundo no-obsoleto mismo año+equipo
        # Uno obsoleto sí convive con uno vigente.
        b1 = self.Budget.search([('team_id', '=', self.team.id)], limit=1)
        b1.state = 'obsoleto'
        self._budget()  # ahora sí

    def test_03_line_date_must_be_first_of_month_in_year(self):
        budget = self._budget()
        with self.assertRaises(ValidationError):
            self._line(budget, date(2040, 6, 15))  # no es día 1
        with self.assertRaises(ValidationError):
            self._line(budget, date(2039, 6, 1))  # fuera del año
        line = self._line(budget, date(2040, 6, 1))
        self.assertTrue(line.id)

    def test_04_uom_must_share_category(self):
        budget = self._budget()
        with self.assertRaises(ValidationError):
            self._line(budget, date(2040, 6, 1), uom=self.uom_kg)  # kg vs producto en m
        line = self._line(budget, date(2040, 6, 1), uom=self.uom_cm)  # cm ~ m: ok
        self.assertTrue(line.id)

    def test_05_display_name_includes_unit(self):
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), uom=self.uom_m)
        self.assertIn('(m)', line.display_name)

    def test_06_real_facturado_moneda_companiacon_balance(self):
        # Factura en MXN (≠ USD compañía): el real usa balance (ya convertido por
        # contabilidad), NO la cifra facial en MXN.
        mxn = self.env.ref('base.MXN')
        mxn.active = True
        self.env['res.currency.rate'].create({
            'currency_id': mxn.id, 'name': date(2040, 6, 1), 'rate': 20.0,
            'company_id': self.company.id})
        self._invoice(date(2040, 6, 10), 5.0, self.uom_m, 400.0, currency=mxn)
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), qty=10.0, amount=150.0)
        # 2000 MXN / 20 = 100 USD (moneda compañía), no 2000.
        self.assertAlmostEqual(line.amount_real, 100.0, places=2)
        self.assertEqual(line.qty_real, 5.0)

    def test_07_real_convierte_unidad_misma_categoria(self):
        # Factura en cm; la línea presupuesta en m → 500 cm = 5 m.
        self._invoice(date(2040, 6, 10), 500.0, self.uom_cm, 1.0)
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), uom=self.uom_m)
        self.assertEqual(line.qty_real, 5.0)

    def test_08_real_refund_resta(self):
        self._invoice(date(2040, 6, 10), 10.0, self.uom_m, 100.0)
        self._invoice(date(2040, 6, 12), 3.0, self.uom_m, 100.0, refund=True)
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), uom=self.uom_m)
        self.assertEqual(line.qty_real, 7.0, "10 facturados - 3 en nota de crédito.")
        self.assertAlmostEqual(line.amount_real, 700.0, places=2)

    def test_09_unidad_de_otra_categoria_excluida_y_contada(self):
        # Una factura del producto con unidad kg (otra categoría) cuenta en importe
        # pero NO en cantidad, y se marca en unconverted_count.
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': self.partner.id,
            'invoice_date': date(2040, 6, 10), 'team_id': self.team.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': self.product.id, 'quantity': 8.0,
                'product_uom_id': self.uom_kg.id, 'price_unit': 50.0,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), uom=self.uom_m)
        self.assertEqual(line.qty_real, 0.0, "kg no convertible a m: excluida de cantidad.")
        self.assertAlmostEqual(line.amount_real, 400.0, places=2, msg="El importe sí cuenta.")
        self.assertEqual(line.unconverted_count, 1)
        self.assertEqual(budget.unconverted_count, 1)

    def test_10_qty_text_por_unidad_sin_mezclar(self):
        budget = self._budget()
        p2 = self.env['product.product'].create({
            'name': 'Fibra PPV', 'type': 'consu', 'uom_id': self.uom_kg.id})
        self._line(budget, date(2040, 6, 1), uom=self.uom_m, qty=1200.0)
        self._line(budget, date(2040, 7, 1), uom=self.uom_m, qty=300.0)
        self._line(budget, date(2040, 6, 1), uom=self.uom_kg, qty=80.0, product=p2)
        text = budget.qty_budget_text
        self.assertIn('1,500 m', text)  # 1200 + 300 m
        self.assertIn('80 kg', text)
        self.assertIn('·', text, "Las unidades se listan por separado, no se suman.")

    def test_11_avg_prices_protege_division_cero(self):
        # El precio ya no se captura: sale de la lista (list_price = 20).
        self.product.list_price = 20.0
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), qty=0.0)
        self.assertEqual(line.avg_price_budget, 0.0)  # sin cantidad → 0, no ZeroDivision
        line2 = self._line(budget, date(2040, 7, 1), qty=50.0)
        self.assertEqual(line2.avg_price_budget, 20.0)  # = precio de lista

    def test_12_locked_when_approved(self):
        manager = self.env['res.users'].create({
            'name': 'MAST PPV', 'login': 'ppv_mgr',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_manager').id])]})
        raso = self.env['res.users'].create({
            'name': 'Raso PPV', 'login': 'ppv_raso',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_user').id])]})
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), qty=10.0)
        budget.with_user(manager).action_send_to_review()
        budget.with_user(manager).action_approve()
        self.assertEqual(budget.state, 'aprobado')
        # Raso no puede editar ni borrar la línea aprobada.
        with self.assertRaises(UserError):
            line.with_user(raso).write({'qty_budget': 99.0})
        with self.assertRaises(UserError):
            line.with_user(raso).unlink()
        # MAST sí (tras regresar a borrador, o directamente por privilegio).
        line.with_user(manager).write({'qty_budget': 42.0})
        self.assertEqual(line.qty_budget, 42.0)

    def test_13_approve_requires_review_manager_and_lines(self):
        manager = self.env['res.users'].create({
            'name': 'MAST PPV 13', 'login': 'ppv_mgr13',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_manager').id])]})
        raso = self.env['res.users'].create({
            'name': 'Raso PPV 2', 'login': 'ppv_raso2',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_user').id])]})
        budget = self._budget()
        # Sin líneas no se puede enviar a revisión.
        with self.assertRaises(UserError):
            budget.with_user(manager).action_send_to_review()
        self._line(budget, date(2040, 6, 1), qty=10.0)
        # Un presupuesto en borrador no se aprueba directo: primero revisado.
        with self.assertRaises(UserError):
            budget.with_user(manager).action_approve()
        budget.with_user(manager).action_send_to_review()
        self.assertEqual(budget.state, 'revisado')
        # Raso (ni Admin de ventas) no puede aprobar: es de la alta dirección.
        with self.assertRaises(UserError):
            budget.with_user(raso).action_approve()

    def test_14_action_revise_conserva_historia(self):
        manager = self.env['res.users'].create({
            'name': 'MAST PPV rev', 'login': 'ppv_mgr_rev',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_manager').id])]})
        budget = self._budget()
        self._line(budget, date(2040, 6, 1), qty=10.0, amount=500.0)
        budget.with_user(manager).action_send_to_review()
        budget.with_user(manager).action_approve()
        action = budget.with_user(manager).action_revise()
        new = self.Budget.browse(action['res_id'])
        self.assertEqual(budget.state, 'obsoleto', "La anterior se conserva obsoleta.")
        self.assertEqual(new.state, 'borrador')
        self.assertEqual(new.revision, 2)
        self.assertEqual(new.year, 2040)
        self.assertEqual(len(new.line_ids), 1, "Copia las líneas.")
        self.assertEqual(new.line_ids.qty_budget, 10.0)

    def test_15_locked_price_and_partner_fields(self):
        # Fix del candado: price_unit_budget/partner_id/customer_code también se
        # bloquean en un presupuesto aprobado para no-MAST.
        raso = self.env['res.users'].create({
            'name': 'Raso PPV precio', 'login': 'ppv_raso_price',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_user').id])]})
        budget = self._budget()
        line = self._line(budget, date(2040, 6, 1), qty=10.0)
        budget.state = 'aprobado'
        for vals in ({'price_unit_budget': 3.0}, {'partner_id': self.env.company.partner_id.id},
                     {'customer_code': 'X'}):
            with self.assertRaises(UserError):
                line.with_user(raso).write(vals)

    def test_16_delete_budget_with_lines(self):
        # Borrar un presupuesto con líneas no truena (los Monetary calculados con
        # currency_field no hacen flush sobre líneas ya eliminadas).
        budget = self._budget()
        self._line(budget, date(2040, 6, 1), qty=10.0)
        self._line(budget, date(2040, 7, 1), qty=20.0)
        line_ids = budget.line_ids.ids
        budget.unlink()
        self.assertFalse(budget.exists())
        self.assertFalse(self.Line.browse(line_ids).exists())

    def test_17_delete_approved_budget(self):
        budget = self._budget()
        self._line(budget, date(2040, 6, 1), qty=10.0)
        budget.state = 'aprobado'
        budget.unlink()
        self.assertFalse(budget.exists())


@tagged('post_install', '-at_install')
class TestSalesBudgetForecast(TransactionCase):
    """Extensión 5.2b — Paso 1: pronóstico semanal por cliente (F-P-A28-13)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado forecast'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.uom_cm = cls.env.ref('uom.product_uom_cm')
        cls.product = cls.env['product.product'].create({
            'name': 'Material forecast', 'default_code': 'SCR31', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.client = cls.env['res.partner'].create(
            {'name': 'Cliente forecast', 'is_company': True})

    def _forecast(self, partner=None):
        return self.Budget.create({
            'year': 2040, 'team_id': self.team.id, 'kind': 'pronostico',
            'partner_id': (partner or self.client).id})

    def _monday(self, ref):
        from datetime import timedelta
        return ref - timedelta(days=ref.weekday())

    def test_01_forecast_line_must_be_monday(self):
        fc = self._forecast()
        # 2040-06-06 no necesariamente es lunes → error si no lo es.
        bad = date(2040, 6, 3)
        monday = self._monday(bad)
        if bad != monday:
            with self.assertRaises(ValidationError):
                self.Line.create({
                    'budget_id': fc.id, 'product_id': self.product.id,
                    'date': bad, 'uom_id': self.uom_m.id})
        line = self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id,
            'date': monday, 'uom_id': self.uom_m.id, 'customer_code': 'SCR31'})
        self.assertEqual(line.date.weekday(), 0)
        self.assertEqual(line.partner_id, self.client, "Cliente heredado de la cabecera.")

    def test_02_forecast_requires_partner(self):
        with self.assertRaises(ValidationError):
            self.Budget.create({
                'year': 2040, 'team_id': self.team.id, 'kind': 'pronostico'})

    def test_03_unique_per_client_and_kind(self):
        self._forecast()
        with self.assertRaises(ValidationError):
            self._forecast()  # mismo cliente-año-equipo-kind
        # Otro cliente sí.
        other = self.env['res.partner'].create(
            {'name': 'Cliente forecast 2', 'is_company': True})
        self._forecast(other)
        # Un presupuesto mensual del mismo equipo/año coexiste (kind distinto).
        self.Budget.create({'year': 2040, 'team_id': self.team.id})

    def test_04_real_weekly_by_commitment_with_conversion(self):
        monday = self._monday(date(2040, 6, 3))
        # Pedido confirmado del cliente, comprometido a un día de esa semana, en cm.
        from datetime import datetime, timedelta
        commit = datetime.combine(monday + timedelta(days=2), datetime.min.time())
        so = self.env['sale.order'].create({
            'partner_id': self.client.id, 'team_id': self.team.id,
            'commitment_date': commit,
            'order_line': [(0, 0, {
                'product_id': self.product.id, 'product_uom_qty': 500.0,
                'product_uom_id': self.uom_cm.id, 'price_unit': 1.0})]})
        so.action_confirm()
        fc = self._forecast()
        line = self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id,
            'date': monday, 'uom_id': self.uom_m.id, 'qty_budget': 4.0})
        # 500 cm comprometidos → 5 m (conversión); es la semana de la línea.
        self.assertEqual(line.qty_real, 5.0)
        # Drill-down: la orden aparece.
        action = line.action_view_week_orders()
        self.assertIn(so.id, action['domain'][0][2])

    def test_05_presupuesto_mode_unchanged_regression(self):
        # Regresión: un presupuesto mensual sigue midiendo FACTURADO, no pedidos.
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        income = self.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': self.client.id,
            'invoice_date': date(2040, 6, 10), 'team_id': self.team.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': self.product.id, 'quantity': 7.0,
                'product_uom_id': self.uom_m.id, 'price_unit': 10.0,
                'account_id': income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        line = self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 5.0})
        self.assertEqual(line.kind, 'presupuesto')
        self.assertEqual(line.qty_real, 7.0, "Sigue siendo facturado, primer día de mes.")
        self.assertEqual(line.date.day, 1)


@tagged('post_install', '-at_install')
class TestSalesBudgetStep2(TransactionCase):
    """Paso 2: grid de captura, comparación, banner de formato e impresión."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado Confección PPV'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Entretela PPV', 'type': 'consu', 'uom_id': cls.uom_m.id})

    def _budget(self):
        return self.Budget.create({'year': 2040, 'team_id': self.team.id})

    def test_01_grid_update_cell_creates_and_increments(self):
        budget = self._budget()
        Line = self.Line.with_context(default_budget_id=budget.id)
        domain = [('budget_id', '=', budget.id),
                  ('product_id', '=', self.product.id),
                  ('date', '>=', '2040-06-01'), ('date', '<', '2040-07-01')]
        # Celda vacía → crea la línea (producto/mes/unidad de venta).
        Line.grid_update_cell(domain, 'qty_budget', 100.0)
        line = self.Line.search([('budget_id', '=', budget.id)])
        self.assertEqual(len(line), 1)
        self.assertEqual(line.date, date(2040, 6, 1))
        self.assertEqual(line.uom_id, self.uom_m)
        self.assertEqual(line.qty_budget, 100.0)
        # Celda existente → suma. (La grid de importes se eliminó: solo cantidades.)
        Line.grid_update_cell(domain, 'qty_budget', 50.0)
        self.assertEqual(line.qty_budget, 150.0)

    def test_02_grid_zero_value_noop(self):
        budget = self._budget()
        Line = self.Line.with_context(default_budget_id=budget.id)
        domain = [('budget_id', '=', budget.id),
                  ('product_id', '=', self.product.id),
                  ('date', '>=', '2040-06-01'), ('date', '<', '2040-07-01')]
        Line.grid_update_cell(domain, 'qty_budget', 0.0)
        self.assertFalse(self.Line.search([('budget_id', '=', budget.id)]))

    def test_03_format_banner_a28_18(self):
        budget = self._budget()
        self.assertIn('F-P-A28-18', budget.sgi_format_banner or '')

    def test_04_grid_and_comparison_actions(self):
        budget = self._budget()
        act_qty = budget.action_open_grid_qty()
        self.assertEqual(act_qty['res_model'], 'sgi.sales.budget.line')
        self.assertTrue(any(v[1] == 'grid' for v in act_qty['views']))
        act_cmp = budget.action_open_comparison()
        self.assertEqual(act_cmp['view_mode'], 'pivot,graph,list')

    def test_05_report_renders(self):
        budget = self._budget()
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id,
            'qty_budget': 1200.0, 'amount_budget': 48000.0})
        html, ttype = self.env['ir.actions.report']._render_qweb_html(
            'quimibond_sgi.action_report_sales_budget', budget.ids)
        self.assertEqual(ttype, 'html')
        self.assertIn('Presupuesto de Ventas', html.decode())
        self.assertIn('F-P-A28-18', html.decode())

    def test_06_pivot_can_aggregate_real(self):
        # El real/pedido están almacenados: el pivot (read_group) puede agregarlos
        # sin "No aggregate function has been provided for the measure".
        budget = self._budget()
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id,
            'qty_budget': 100.0, 'amount_budget': 5000.0})
        groups = self.Line._read_group(
            [('budget_id', '=', budget.id)], groupby=['product_id'],
            aggregates=['amount_budget:sum', 'amount_real:sum',
                        'amount_ordered:sum', 'qty_real:sum'])
        self.assertTrue(groups)


@tagged('post_install', '-at_install')
class TestSalesBudgetStep3(TransactionCase):
    """Paso 3: conexión al SGI — KPI VE-02 y aviso de cierre de mes."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Indicator = cls.env['sgi.indicator']
        cls.Cron = cls.env['sgi.cron']
        cls.company = cls.env.company
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Producto VE02', 'type': 'consu', 'uom_id': cls.uom_m.id})
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        cls.partner = cls.env['res.partner'].create({'name': 'Cliente VE02'})

    def _team(self, name):
        return self.env['crm.team'].create({'name': name})

    def _approved_budget(self, team, when, amount):
        # El importe ya no se captura: sale de la lista. Lista fija determinista
        # (cliente) para que qty(100) × precio = amount (presupuesto = importe).
        company_ccy = self.env.company.currency_id
        pl = self.env['product.pricelist'].create(
            {'name': 'PL VE02 %s' % team.name, 'currency_id': company_ccy.id})
        self.env['product.pricelist.item'].create({
            'pricelist_id': pl.id, 'applied_on': '1_product',
            'product_tmpl_id': self.product.product_tmpl_id.id,
            'compute_price': 'fixed', 'fixed_price': amount / 100.0})
        partner = self.env['res.partner'].create(
            {'name': 'Cliente %s' % team.name, 'is_company': True})
        partner.property_product_pricelist = pl
        budget = self.Budget.create({'year': when.year, 'team_id': team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': when, 'uom_id': self.uom_m.id, 'qty_budget': 100.0,
            'partner_id': partner.id})
        budget.state = 'aprobado'
        return budget

    def _invoice(self, team, when, amount, refund=False):
        move = self.env['account.move'].create({
            'move_type': 'out_refund' if refund else 'out_invoice',
            'partner_id': self.partner.id, 'invoice_date': when,
            'team_id': team.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': self.product.id, 'quantity': 1,
                'product_uom_id': self.uom_m.id, 'price_unit': amount,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    def test_01_ve02_lee_presupuesto_aprobado(self):
        team = self._team('Mercado VE02 A')
        self._approved_budget(team, date(2040, 6, 1), 1000.0)
        # Facturación neta junio 2040 = 1000 - 200 = 800 → 80% del presupuesto.
        self._invoice(team, date(2040, 6, 10), 1000.0)
        self._invoice(team, date(2040, 6, 12), 200.0, refund=True)
        ind = self.Indicator.create({
            'code': 'VE02-T', 'name': 'Cumpl. presupuesto',
            'calc_mode': 'presupuesto_ventas', 'direction': 'higher_better'})
        value = ind._calc_presupuesto_ventas(date(2040, 6, 1), date(2040, 6, 30))
        self.assertEqual(value, 80.0)
        # Sin nota: hay presupuesto aprobado.
        self.assertFalse(ind._note_presupuesto_ventas(date(2040, 6, 1), date(2040, 6, 30)))

    def test_02_ve02_fallback_al_parametro_con_nota(self):
        # Sin presupuesto aprobado del año → usa el parámetro de Ajustes + nota.
        self.env['ir.config_parameter'].sudo().set_param(
            'quimibond_sgi.monthly_sales_budget', '1000')
        team = self._team('Mercado VE02 B')
        self._invoice(team, date(2040, 6, 10), 800.0)
        ind = self.Indicator.create({
            'code': 'VE02-F', 'name': 'Cumpl. presupuesto fb',
            'calc_mode': 'presupuesto_ventas', 'direction': 'higher_better'})
        value = ind._calc_presupuesto_ventas(date(2040, 6, 1), date(2040, 6, 30))
        self.assertEqual(value, 80.0)
        self.assertIn('Ajustes',
                      ind._note_presupuesto_ventas(date(2040, 6, 1), date(2040, 6, 30)))

    def test_03_ve02_evidencia_son_las_lineas(self):
        team = self._team('Mercado VE02 C')
        budget = self._approved_budget(team, date(2040, 6, 1), 1000.0)
        ind = self.Indicator.create({
            'code': 'VE02-E', 'name': 'Cumpl. presupuesto ev',
            'calc_mode': 'presupuesto_ventas', 'direction': 'higher_better'})
        measure = self.env['sgi.indicator.measure'].create({
            'indicator_id': ind.id, 'period_date': date(2040, 6, 1),
            'value': 80.0, 'state': 'capturado'})
        action = measure.action_view_evidence()
        self.assertEqual(action['res_model'], 'sgi.sales.budget.line')
        records = self.Line.search(action['domain'])
        self.assertEqual(records, budget.line_ids)

    def test_04_cierre_de_mes_avisa_bajo_umbral(self):
        team = self._team('Mercado cierre bajo')
        team.user_id = self.env.user.id
        self._approved_budget(team, date(2040, 6, 1), 1000.0)
        self._invoice(team, date(2040, 6, 10), 500.0)  # 50% < 80%
        self.Cron._sgi_sales_budget_month_close(date(2040, 6, 1), date(2040, 6, 30))
        acts = self.env['mail.activity'].search([
            ('res_model', '=', 'sgi.sales.budget'),
            ('user_id', '=', self.env.user.id)])
        self.assertTrue(acts, "Equipo bajo umbral debe recibir aviso de cierre.")

    def test_05_cierre_de_mes_no_avisa_si_cumple(self):
        team = self._team('Mercado cierre ok')
        team.user_id = self.env.user.id
        self._approved_budget(team, date(2040, 6, 1), 1000.0)
        self._invoice(team, date(2040, 6, 10), 950.0)  # 95% >= 80%
        before = self.env['mail.activity'].search_count([
            ('res_model', '=', 'sgi.sales.budget')])
        self.Cron._sgi_sales_budget_month_close(date(2040, 6, 1), date(2040, 6, 30))
        after = self.env['mail.activity'].search_count([
            ('res_model', '=', 'sgi.sales.budget')])
        self.assertEqual(before, after, "Equipo que cumple no genera aviso.")


@tagged('post_install', '-at_install')
class TestSalesBudgetClient(TransactionCase):
    """Adición 5.2: dimensión cliente (opcional, sin doble conteo)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado cliente PPV'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Tela cliente PPV', 'type': 'consu', 'uom_id': cls.uom_m.id})
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        # Empresa comercial C con un contacto hijo CC, y otra empresa D.
        cls.company_c = cls.env['res.partner'].create(
            {'name': 'Cliente C', 'is_company': True})
        cls.contact_cc = cls.env['res.partner'].create(
            {'name': 'Contacto CC', 'parent_id': cls.company_c.id})
        cls.company_d = cls.env['res.partner'].create(
            {'name': 'Cliente D', 'is_company': True})

    def _budget(self):
        return self.Budget.create({'year': 2040, 'team_id': self.team.id})

    def _line(self, budget, partner=None, amount=0.0, product=None):
        return self.Line.create({
            'budget_id': budget.id, 'product_id': (product or self.product).id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id,
            'partner_id': partner.id if partner else False,
            'qty_budget': 10.0, 'amount_budget': amount})

    def _invoice(self, partner, amount):
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': partner.id,
            'invoice_date': date(2040, 6, 10), 'team_id': self.team.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': self.product.id, 'quantity': 1,
                'product_uom_id': self.uom_m.id, 'price_unit': amount,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    def test_01_double_count_blocked(self):
        budget = self._budget()
        self._line(budget, partner=None, amount=1000.0)  # global
        # Mezclar el mismo producto con cliente → bloqueado.
        with self.assertRaises(ValidationError):
            self._line(budget, partner=self.company_c, amount=500.0)

    def test_02_double_count_blocked_reverse(self):
        budget = self._budget()
        self._line(budget, partner=self.company_c, amount=500.0)  # por cliente
        with self.assertRaises(ValidationError):
            self._line(budget, partner=None, amount=1000.0)  # ahora global → bloqueado

    def test_03_two_clients_ok(self):
        budget = self._budget()
        self._line(budget, partner=self.company_c, amount=500.0)
        line_d = self._line(budget, partner=self.company_d, amount=300.0)
        self.assertTrue(line_d.id, "Dos clientes distintos para el mismo producto: OK.")

    def test_04_unique_global_per_product_month(self):
        budget = self._budget()
        self._line(budget, partner=None, amount=1000.0)
        # Segundo global mismo producto+mes → viola el índice único.
        with self.assertRaises(Exception):
            with self.env.cr.savepoint():
                self._line(budget, partner=None, amount=200.0)

    def test_05_real_filtered_by_commercial_partner(self):
        # Factura al CONTACTO hijo de C, y otra a D. La línea por cliente C debe
        # tomar solo la de C (vía commercial_partner_id del documento).
        self._invoice(self.contact_cc, 500.0)
        self._invoice(self.company_d, 300.0)
        budget = self._budget()
        line = self._line(budget, partner=self.company_c, amount=400.0)
        self.assertAlmostEqual(line.amount_real, 500.0, places=2)

    def test_06_amount_real_unbudgeted(self):
        # Equipo factura 800 (500 a C, 300 a D); se presupuesta solo C → 300 sin
        # presupuestar.
        self._invoice(self.contact_cc, 500.0)
        self._invoice(self.company_d, 300.0)
        budget = self._budget()
        self._line(budget, partner=self.company_c, amount=400.0)
        self.assertAlmostEqual(budget.amount_real_unbudgeted, 300.0, places=2)

    def test_07_global_line_takes_all_clients(self):
        # Sin dimensión cliente, el real es de todo el equipo (C + D).
        self._invoice(self.contact_cc, 500.0)
        self._invoice(self.company_d, 300.0)
        budget = self._budget()
        line = self._line(budget, partner=None, amount=1000.0)
        self.assertAlmostEqual(line.amount_real, 800.0, places=2)
        self.assertAlmostEqual(budget.amount_real_unbudgeted, 0.0, places=2)


@tagged('post_install', '-at_install')
class TestSalesBudgetOnlyQty(TransactionCase):
    """5.2d Paso 1: solo cantidades — el precio manda la lista, en doble moneda."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Param = cls.env['ir.config_parameter'].sudo()
        cls.team = cls.env['crm.team'].create({'name': 'Mercado precio PPV'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.usd = cls.env.ref('base.USD')  # moneda de la compañía en esta BD
        cls.mxn = cls.env.ref('base.MXN')
        cls.mxn.active = True
        cls.product = cls.env['product.product'].create({
            'name': 'Tela precio PPV', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 100.0})
        cls.budget = cls.Budget.create({'year': 2040, 'team_id': cls.team.id})

    def _pricelist(self, currency, price):
        pl = self.env['product.pricelist'].create({
            'name': 'Lista %s PPV' % currency.name, 'currency_id': currency.id})
        self.env['product.pricelist.item'].create({
            'pricelist_id': pl.id, 'applied_on': '1_product',
            'product_tmpl_id': self.product.product_tmpl_id.id,
            'compute_price': 'fixed', 'fixed_price': price})
        return pl

    def _partner_with_list(self, pricelist, name='Cliente lista PPV'):
        partner = self.env['res.partner'].create({'name': name, 'is_company': True})
        partner.property_product_pricelist = pricelist
        return partner

    def _line(self, budget, partner=None, qty=10.0, when=None, product=None):
        return self.Line.create({
            'budget_id': budget.id, 'product_id': (product or self.product).id,
            'uom_id': self.uom_m.id, 'qty_budget': qty,
            'date': when or date(2040, 6, 1),
            'partner_id': partner.id if partner else False})

    def test_01_price_from_client_list_and_readonly(self):
        # Precio SIEMPRE de la lista del cliente (USD = compañía): 37.63.
        partner = self._partner_with_list(self._pricelist(self.usd, 37.63))
        line = self._line(self.budget, partner)
        self.assertAlmostEqual(line.price_unit_budget, 37.63, places=2)
        self.assertAlmostEqual(line.amount_budget, 376.3, places=2)  # 10 × 37.63
        self.assertIn("Lista", line.price_source)
        # La lista es la fuente de verdad: aunque se fuerce otro valor, el refresh
        # (botón/cron) lo devuelve al de la lista (en borrador). En la vista es
        # readonly; aquí probamos que el precio SIEMPRE vuelve a la lista.
        line.write({'price_unit_budget': 999.0})
        self.budget.action_refresh_actuals()
        self.assertAlmostEqual(line.price_unit_budget, 37.63, places=2)

    def test_02_dual_currency_usd_client(self):
        # Cliente con lista en MXN (≠ USD compañía), tipo presupuestal 10:
        # precio divisa = 350 MXN; precio compañía = 35 USD.
        self.Param.set_param('quimibond_sgi.budget_planning_rate', '10')
        partner = self._partner_with_list(self._pricelist(self.mxn, 350.0))
        line = self._line(self.budget, partner)
        self.assertEqual(line.list_currency_id, self.mxn)
        self.assertAlmostEqual(line.price_unit_currency, 350.0, places=2)
        self.assertAlmostEqual(line.amount_currency, 3500.0, places=2)  # 10 × 350
        self.assertAlmostEqual(line.price_unit_budget, 35.0, places=2)
        self.assertAlmostEqual(line.amount_budget, 350.0, places=2)  # 10 × 35

    def test_03_mxn_client_no_currency_columns(self):
        # Cliente con lista en moneda de la compañía (USD): sin divisa aparte.
        partner = self._partner_with_list(self._pricelist(self.usd, 20.0))
        line = self._line(self.budget, partner)
        self.assertEqual(line.list_currency_id, self.usd)
        self.assertEqual(line.list_currency_id, line.currency_id,
                         "Lista en moneda compañía: las columnas de divisa se ocultan.")

    def test_04_currency_text_mixed_clients(self):
        self.Param.set_param('quimibond_sgi.budget_planning_rate', '10')
        budget = self.Budget.create({'year': 2041, 'team_id': self.team.id})
        p_usd = self._partner_with_list(self._pricelist(self.usd, 20.0), 'C USD')
        # Segundo producto para el cliente MXN (evita esquema mixto por producto).
        prod2 = self.env['product.product'].create(
            {'name': 'Tela 2 PPV', 'type': 'consu', 'uom_id': self.uom_m.id})
        pl_mxn = self.env['product.pricelist'].create(
            {'name': 'PL MXN', 'currency_id': self.mxn.id})
        self.env['product.pricelist.item'].create({
            'pricelist_id': pl_mxn.id, 'applied_on': '1_product',
            'product_tmpl_id': prod2.product_tmpl_id.id,
            'compute_price': 'fixed', 'fixed_price': 350.0})
        p_mxn = self.env['res.partner'].create({'name': 'C MXN', 'is_company': True})
        p_mxn.property_product_pricelist = pl_mxn
        self._line(budget, p_usd, qty=10.0, when=date(2041, 6, 1))
        self._line(budget, p_mxn, qty=10.0, when=date(2041, 6, 1), product=prod2)
        text = budget.amount_currency_text
        self.assertIn('MXN', text)  # divisa por moneda
        self.assertIn('Total compañía', text)  # y el único total global en pesos

    def test_05_no_list_price_banner(self):
        # Producto sin precio en ninguna lista (list_price 0) → sin precio de lista.
        prod0 = self.env['product.product'].create(
            {'name': 'Sin precio PPV', 'type': 'consu', 'uom_id': self.uom_m.id,
             'list_price': 0.0})
        budget = self.Budget.create({'year': 2042, 'team_id': self.team.id})
        line = self._line(budget, qty=5.0, when=date(2042, 6, 1), product=prod0)
        self.assertFalse(line.has_list_price)
        self.assertEqual(budget.no_price_count, 1)

    def test_06_price_frozen_on_approve(self):
        partner = self._partner_with_list(self._pricelist(self.usd, 50.0))
        budget = self.Budget.create({'year': 2043, 'team_id': self.team.id})
        line = self._line(budget, partner, when=date(2043, 6, 1))
        self.assertAlmostEqual(line.price_unit_budget, 50.0, places=2)
        budget.state = 'aprobado'
        # Cambia la lista y refresca: el aprobado NO se toca (congelado).
        partner.property_product_pricelist.item_ids.fixed_price = 80.0
        budget.action_refresh_actuals()
        self.assertAlmostEqual(line.price_unit_budget, 50.0, places=2,
                               msg="Aprobado: precio congelado.")


@tagged('post_install', '-at_install')
class TestSalesBudgetImport(TransactionCase):
    """Adición 5.2: asistente de importación del Excel F-P-A28-18."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Wizard = cls.env['sgi.sales.budget.import']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado import PPV'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.uom_kg = cls.env.ref('uom.product_uom_kgm')
        cls.tela = cls.env['product.product'].create({
            'name': 'Tela import', 'default_code': 'TELA-1', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 30.0})
        cls.fibra = cls.env['product.product'].create({
            'name': 'Fibra import', 'default_code': 'FIBRA-1', 'type': 'consu',
            'uom_id': cls.uom_kg.id})
        # Lista presupuestal para valuar las líneas globales del import (sin
        # cliente): TELA-1 a 30 (P5 — ya no se toma una lista arbitraria).
        cls.budget_pl = cls.env['product.pricelist'].create(
            {'name': 'Presupuestal import PPV'})
        cls.env['product.pricelist.item'].create({
            'pricelist_id': cls.budget_pl.id, 'applied_on': '1_product',
            'product_tmpl_id': cls.tela.product_tmpl_id.id,
            'compute_price': 'fixed', 'fixed_price': 30.0})
        cls.env['ir.config_parameter'].sudo().set_param(
            'quimibond_sgi.budget_pricelist_id', cls.budget_pl.id)

    def _xlsx(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Presupuesto'
        ws.append(['PRODUCTO', 'UNIDAD', 'ENERO m', 'ENERO $',
                   'FEBRERO m', 'FEBRERO $', 'MARZO m', 'MARZO $'])
        ws.append(['TELA-1', 'm', 100, 5000, 120, 6000, 0, 0])
        ws.append(['FIBRA-1', 'kg', 50, 2500, 0, 0, 0, 0])
        ws.append(['NOEXISTE', 'm', 10, 500, 0, 0, 0, 0])
        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue())

    def _wizard(self, budget, mode='replace'):
        return self.Wizard.create({
            'budget_id': budget.id, 'file': self._xlsx(),
            'filename': 'ppv.xlsx', 'sheet_name': 'Presupuesto',
            'conflict_mode': mode})

    def test_01_import_matrix(self):
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        wiz = self._wizard(budget)
        wiz.action_import()
        # TELA-1: enero + febrero (2 líneas); FIBRA-1: enero (1). Total 3.
        self.assertEqual(len(budget.line_ids), 3)
        tela_ene = budget.line_ids.filtered(
            lambda l: l.product_id == self.tela and l.date == date(2040, 1, 1))
        self.assertEqual(tela_ene.qty_budget, 100.0)
        # El $ del Excel (5000) se IGNORA: el importe sale de la lista
        # (100 × list_price 30 = 3000), y el resultado lo nota.
        self.assertAlmostEqual(tela_ene.amount_budget, 3000.0, places=2)
        self.assertIn('importes se calculan de la lista', wiz.result)
        self.assertEqual(tela_ene.uom_id, self.uom_m)
        fibra = budget.line_ids.filtered(lambda l: l.product_id == self.fibra)
        self.assertEqual(fibra.uom_id, self.uom_kg, "Unidad kg de la columna UNIDAD.")
        self.assertEqual(fibra.qty_budget, 50.0)

    def test_02_unmatched_reported_not_fatal(self):
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        wiz = self._wizard(budget)
        wiz.action_import()
        self.assertIn('NOEXISTE', wiz.result)
        self.assertIn('sin match', wiz.result)
        # A pesar del no-match, las líneas buenas se importaron.
        self.assertEqual(len(budget.line_ids), 3)

    def test_03_reimport_replace_no_duplicate(self):
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        self._wizard(budget).action_import()
        self.assertEqual(len(budget.line_ids), 3)
        # Reimportar en modo reemplazar no duplica.
        self._wizard(budget).action_import()
        self.assertEqual(len(budget.line_ids), 3)

    def test_04_no_header_is_structural_error(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['COSA', 'OTRA'])  # sin 'PRODUCTO'
        ws.append(['x', 'y'])
        buf = io.BytesIO()
        wb.save(buf)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        wiz = self.Wizard.create({
            'budget_id': budget.id, 'file': base64.b64encode(buf.getvalue()),
            'sheet_name': 'Sheet'})
        with self.assertRaises(UserError):
            wiz.action_import()
        self.assertFalse(budget.line_ids, "Error estructural: se revierte todo.")

    def test_05_approved_budget_blocks_import(self):
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        self.env['sgi.sales.budget.line'].create({
            'budget_id': budget.id, 'product_id': self.tela.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 1.0})
        budget.state = 'aprobado'
        with self.assertRaises(UserError):
            budget.action_open_import()


@tagged('post_install', '-at_install')
class TestSalesBudgetForecastImport(TransactionCase):
    """Extensión 5.2b — Paso 2: importación e impresión del forecast."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Wizard = cls.env['sgi.sales.budget.import']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado fc import'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Material SCR31', 'default_code': 'SCR31', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.client = cls.env['res.partner'].create(
            {'name': 'Cliente forecast import', 'is_company': True})

    def _forecast(self):
        return self.Budget.create({
            'year': 2040, 'team_id': self.team.id, 'kind': 'pronostico',
            'partner_id': self.client.id})

    def _forecast_xlsx(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cliente A'
        ws.append(['SEMANA', '', 23, 24, 25])
        # Dos bloques del mismo producto (oleadas de PO) → se suman por semana.
        ws.append(['SCR31', 'SCR31', '16,000', 1000, 0])  # coma de miles
        ws.append(['SCR31', 'SCR31', 4000, 0, 500])
        ws.append(['PO', '', 0, 0, 0])       # fila PO → se ignora
        ws.append(['TOTAL', '', 20000, 1000, 500])  # fila TOTAL → se ignora
        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue())

    def test_01_import_forecast_layout(self):
        fc = self._forecast()
        wiz = self.Wizard.create({
            'budget_id': fc.id, 'file': self._forecast_xlsx(),
            'sheet_name': 'Cliente A', 'conflict_mode': 'replace'})
        wiz.action_import()
        # 3 semanas (23, 24, 25) para SCR31.
        self.assertEqual(len(fc.line_ids), 3)
        by_week = {l.date: l for l in fc.line_ids}
        wk23 = self.Wizard._week_monday(2040, 23)
        line23 = by_week[wk23]
        self.assertEqual(line23.qty_budget, 20000.0,
                         "16,000 (coma) + 4,000 sumados por semana.")
        self.assertEqual(line23.customer_code, 'SCR31')
        self.assertEqual(line23.partner_id, self.client)
        self.assertEqual(line23.date.weekday(), 0, "Lunes de la semana.")

    def test_02_forecast_report_renders(self):
        fc = self._forecast()
        self.env['sgi.sales.budget.line'].create({
            'budget_id': fc.id, 'product_id': self.product.id,
            'date': self.Wizard._week_monday(2040, 23), 'uom_id': self.uom_m.id,
            'customer_code': 'SCR31', 'qty_budget': 500.0})
        html, ttype = self.env['ir.actions.report']._render_qweb_html(
            'quimibond_sgi.action_report_sales_budget', fc.ids)
        self.assertIn('Pronóstico de Ventas', html.decode())
        self.assertIn('F-P-A28-13', html.decode())
        self.assertIn('SCR31', html.decode())

    def test_03_format_banner_by_kind(self):
        fc = self._forecast()
        self.assertIn('F-P-A28-13', fc.sgi_format_banner or '')
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        self.assertIn('F-P-A28-18', budget.sgi_format_banner or '')

    def test_04_monthly_import_still_works_regression(self):
        # Regresión: un presupuesto mensual sigue usando el layout F-P-A28-18.
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Presupuesto'
        ws.append(['PRODUCTO', 'ENERO m', 'ENERO $'])
        ws.append(['SCR31', 100, 5000])
        buf = io.BytesIO()
        wb.save(buf)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        wiz = self.Wizard.create({
            'budget_id': budget.id, 'file': base64.b64encode(buf.getvalue()),
            'sheet_name': 'Presupuesto'})
        wiz.action_import()
        self.assertEqual(len(budget.line_ids), 1)
        self.assertEqual(budget.line_ids.date, date(2040, 1, 1))
        self.assertEqual(budget.line_ids.qty_budget, 100.0)


@tagged('post_install', '-at_install')
class TestSalesBudgetNetDemand(TransactionCase):
    """Extensión 5.2c: consumo de pronóstico + demanda neta al MPS."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado neta'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.uom_cm = cls.env.ref('uom.product_uom_cm')
        cls.product = cls.env['product.product'].create({
            'name': 'Material neta', 'default_code': 'NET1', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.client = cls.env['res.partner'].create(
            {'name': 'Cliente neta', 'is_company': True})

    def _forecast(self):
        return self.Budget.create({
            'year': 2040, 'team_id': self.team.id, 'kind': 'pronostico',
            'partner_id': self.client.id})

    def _monday(self, ref):
        from datetime import timedelta
        return ref - timedelta(days=ref.weekday())

    def _order(self, monday, qty, uom=None):
        from datetime import datetime, timedelta
        commit = datetime.combine(monday + timedelta(days=1), datetime.min.time())
        so = self.env['sale.order'].create({
            'partner_id': self.client.id, 'team_id': self.team.id,
            'commitment_date': commit,
            'order_line': [(0, 0, {
                'product_id': self.product.id, 'product_uom_qty': qty,
                'product_uom_id': (uom or self.uom_m).id, 'price_unit': 1.0})]})
        so.action_confirm()
        return so

    def test_01_net_demand_order_bigger(self):
        monday = self._monday(date(2040, 6, 3))
        self._order(monday, 150.0)  # comprometido > pronóstico
        fc = self._forecast()
        line = self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id, 'date': monday,
            'uom_id': self.uom_m.id, 'qty_budget': 100.0})
        self.assertEqual(line.qty_real, 150.0)
        self.assertEqual(line.qty_net_demand, 150.0, "Manda el pedido si supera.")

    def test_02_net_demand_order_smaller(self):
        monday = self._monday(date(2040, 6, 10))
        self._order(monday, 60.0)  # comprometido < pronóstico
        fc = self._forecast()
        line = self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id, 'date': monday,
            'uom_id': self.uom_m.id, 'qty_budget': 100.0})
        self.assertEqual(line.qty_net_demand, 100.0, "Manda el pronóstico si es mayor.")

    def test_03_preload_idempotent_no_overwrite(self):
        m1 = self._monday(date(2040, 6, 3))
        m2 = self._monday(date(2040, 6, 17))
        self._order(m1, 80.0)
        self._order(m2, 40.0)
        fc = self._forecast()
        # Captura previa en m1 (no debe pisarse).
        self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id, 'date': m1,
            'uom_id': self.uom_m.id, 'qty_budget': 999.0})
        fc.action_preload_from_orders()
        # m1 intacto; m2 creado con lo comprometido.
        line_m1 = fc.line_ids.filtered(lambda l: l.date == m1)
        line_m2 = fc.line_ids.filtered(lambda l: l.date == m2)
        self.assertEqual(line_m1.qty_budget, 999.0, "No pisa la captura.")
        self.assertEqual(line_m2.qty_budget, 40.0, "Precarga lo comprometido.")
        # Idempotente: segunda corrida no duplica.
        before = len(fc.line_ids)
        fc.action_preload_from_orders()
        self.assertEqual(len(fc.line_ids), before)

    def test_04_mps_available_reflects_module(self):
        fc = self._forecast()
        self.assertEqual(fc.sgi_mps_available,
                         'mrp.production.schedule' in self.env)

    def test_05_send_net_demand_to_mps_idempotent(self):
        if 'mrp.production.schedule' not in self.env:
            self.skipTest("mrp_mps no está instalado en esta BD.")
        monday = self._monday(date(2040, 6, 3))
        fc = self._forecast()
        # Línea en cm: 500 cm de demanda neta → 5 m (unidad del producto) al MPS.
        self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id, 'date': monday,
            'uom_id': self.uom_cm.id, 'qty_budget': 500.0})
        # El pronóstico es documento vivo: su estado terminal es 'revisado'.
        fc.state = 'revisado'
        fc.action_send_to_mps()
        wh = self.env['stock.warehouse'].search(
            [('company_id', '=', fc.company_id.id)], limit=1)
        sched = self.env['mrp.production.schedule'].search([
            ('product_id', '=', self.product.id), ('warehouse_id', '=', wh.id)])
        self.assertTrue(sched)
        forecasts = sched.forecast_ids.filtered(lambda f: f.date == monday)
        self.assertEqual(len(forecasts), 1)
        self.assertAlmostEqual(forecasts.forecast_qty, 5.0, places=2,
                               msg="Convertido a la unidad del producto (m).")
        # Re-envío tras revisión: actualiza sin duplicar.
        fc.action_send_to_mps()
        forecasts = sched.forecast_ids.filtered(lambda f: f.date == monday)
        self.assertEqual(len(forecasts), 1, "No duplica en el re-envío.")

    def test_06_no_sale_orders_created(self):
        # PROHIBIDO crear pedidos desde el pronóstico: la precarga LEE pedidos,
        # nunca los crea.
        monday = self._monday(date(2040, 6, 3))
        fc = self._forecast()
        self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id, 'date': monday,
            'uom_id': self.uom_m.id, 'qty_budget': 100.0})
        before = self.env['sale.order'].search_count([('partner_id', '=', self.client.id)])
        fc.action_preload_from_orders()
        after = self.env['sale.order'].search_count([('partner_id', '=', self.client.id)])
        self.assertEqual(before, after, "El pronóstico no crea pedidos de venta.")


@tagged('post_install', '-at_install')
class TestSalesBudgetTemplate(TransactionCase):
    """5.2d Paso 2: plantilla descargable, roundtrip, menús por tipo."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado plantilla'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        cls.prod_a = cls.env['product.product'].create({
            'name': 'Prod A tpl', 'default_code': 'PA-TPL', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.prod_b = cls.env['product.product'].create({
            'name': 'Prod B tpl', 'default_code': 'PB-TPL', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.cli1 = cls.env['res.partner'].create(
            {'name': 'AAA Cliente uno', 'is_company': True})
        cls.cli2 = cls.env['res.partner'].create(
            {'name': 'BBB Cliente dos', 'is_company': True})

    def _invoice(self, client, product):
        # Fecha reciente (dentro de los 24 meses) para que entre al historial.
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': client.id,
            'invoice_date': date.today(), 'team_id': self.team.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': product.id, 'quantity': 1,
                'product_uom_id': self.uom_m.id, 'price_unit': 10.0,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    def _read_template(self, budget):
        import base64
        import io
        import openpyxl
        action = budget.action_download_template()
        att_id = int(action['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(att.datas)))
        return wb.active, att

    def _data_rows(self, ws):
        """[(producto, cliente)] de las filas con producto (header en fila 2)."""
        rows = []
        for r in range(3, ws.max_row + 1):
            prod = ws.cell(r, 1).value
            if prod:
                rows.append((prod, ws.cell(r, 2).value))
        return rows

    def test_01_template_has_client_product_pairs(self):
        # 2 clientes × 2 productos facturados → 4 pares (cliente × producto).
        self._invoice(self.cli1, self.prod_a)
        self._invoice(self.cli1, self.prod_b)
        self._invoice(self.cli2, self.prod_a)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        ws, _ = self._read_template(budget)
        # Header en la fila 2 (fila 1 = instrucción).
        self.assertEqual(ws.cell(2, 1).value, 'PRODUCTO')
        self.assertEqual(ws.cell(2, 2).value, 'CLIENTE')
        pairs = self._data_rows(ws)
        self.assertIn(('PA-TPL', 'AAA Cliente uno'), pairs)
        self.assertIn(('PB-TPL', 'AAA Cliente uno'), pairs)
        self.assertIn(('PA-TPL', 'BBB Cliente dos'), pairs)
        self.assertNotIn(('PB-TPL', 'BBB Cliente dos'), pairs,
                         "Par no facturado: no aparece.")
        # Ordenado por cliente y luego producto.
        self.assertEqual(pairs[0], ('PA-TPL', 'AAA Cliente uno'))

    def test_02_roundtrip_download_fill_import(self):
        self._invoice(self.cli1, self.prod_a)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        ws, att = self._read_template(budget)
        # Fila del par (PA-TPL, cli1); ENERO m está en la columna 4.
        row = next(r for r in range(3, ws.max_row + 1)
                   if ws.cell(r, 1).value == 'PA-TPL')
        ws.cell(row, 4).value = 250
        import base64
        import io
        buf = io.BytesIO()
        ws.parent.save(buf)
        wiz = self.env['sgi.sales.budget.import'].create({
            'budget_id': budget.id, 'file': base64.b64encode(buf.getvalue()),
            'sheet_name': ws.title})
        wiz.action_import()
        line = budget.line_ids.filtered(lambda l: l.product_id == self.prod_a)
        self.assertEqual(len(line), 1)
        self.assertEqual(line.date, date(2040, 1, 1))
        self.assertEqual(line.qty_budget, 250.0)
        self.assertEqual(line.partner_id, self.cli1, "Línea por cliente del par.")

    def test_02b_empty_rows_do_not_import(self):
        # Plantilla grande, ninguna cantidad → cero líneas (filas vacías se ignoran).
        self._invoice(self.cli1, self.prod_a)
        self._invoice(self.cli2, self.prod_b)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        ws, att = self._read_template(budget)
        import base64
        import io
        buf = io.BytesIO()
        ws.parent.save(buf)
        wiz = self.env['sgi.sales.budget.import'].create({
            'budget_id': budget.id, 'file': base64.b64encode(buf.getvalue()),
            'sheet_name': ws.title})
        wiz.action_import()
        self.assertFalse(budget.line_ids, "Filas sin cantidad no generan líneas.")

    def test_02c_row_without_client_makes_global_line(self):
        self._invoice(self.cli1, self.prod_a)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        ws, att = self._read_template(budget)
        row = next(r for r in range(3, ws.max_row + 1)
                   if ws.cell(r, 1).value == 'PA-TPL')
        ws.cell(row, 2).value = None   # borra el cliente → global
        ws.cell(row, 4).value = 120
        import base64
        import io
        buf = io.BytesIO()
        ws.parent.save(buf)
        wiz = self.env['sgi.sales.budget.import'].create({
            'budget_id': budget.id, 'file': base64.b64encode(buf.getvalue()),
            'sheet_name': ws.title})
        wiz.action_import()
        line = budget.line_ids.filtered(lambda l: l.product_id == self.prod_a)
        self.assertEqual(len(line), 1)
        self.assertFalse(line.partner_id, "Sin cliente en la fila → línea global.")
        self.assertEqual(line.qty_budget, 120.0)

    def test_03_actions_filter_by_kind(self):
        b = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        f = self.Budget.create({'year': 2040, 'team_id': self.team.id,
                                'kind': 'pronostico', 'partner_id': self.cli1.id})
        act_b = self.env.ref('quimibond_sgi.sgi_sales_budget_action')
        act_f = self.env.ref('quimibond_sgi.sgi_sales_forecast_action')
        budgets = self.Budget.search(
            [('team_id', '=', self.team.id)] + list(eval(act_b.domain)))
        forecasts = self.Budget.search(
            [('team_id', '=', self.team.id)] + list(eval(act_f.domain)))
        self.assertIn(b, budgets)
        self.assertNotIn(f, budgets)
        self.assertIn(f, forecasts)
        self.assertNotIn(b, forecasts)

    def test_04_template_only_on_draft(self):
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        budget.state = 'aprobado'
        with self.assertRaises(UserError):
            budget.action_download_template()


@tagged('post_install', '-at_install')
class TestSalesBudgetAnalysis(TransactionCase):
    """5.2d Paso 3: drill-down de facturas, curva y top-5 brechas del cierre."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Cron = cls.env['sgi.cron']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado analisis'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        cls.partner = cls.env['res.partner'].create({'name': 'Cliente analisis'})
        cls.pa = cls.env['product.product'].create({
            'name': 'Prod A an', 'default_code': 'PA-AN', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 10.0})
        cls.pb = cls.env['product.product'].create({
            'name': 'Prod B an', 'default_code': 'PB-AN', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 10.0})
        _seed_budget_pricelist(cls.env, 10.0)  # líneas globales a 10 (P5)

    def _invoice(self, product, when, qty, price):
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': self.partner.id,
            'invoice_date': when, 'team_id': self.team.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': product.id, 'quantity': qty,
                'product_uom_id': self.uom_m.id, 'price_unit': price,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    def test_01_view_month_invoices(self):
        inv = self._invoice(self.pa, date(2040, 6, 10), 5.0, 10.0)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        line = self.Line.create({
            'budget_id': budget.id, 'product_id': self.pa.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        action = line.action_view_month_invoices()
        self.assertEqual(action['res_model'], 'account.move')
        moves = self.env['account.move'].search(action['domain'])
        self.assertIn(inv, moves)

    def test_02_cumulative_action(self):
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        action = budget.action_open_cumulative()
        self.assertEqual(action['res_model'], 'sgi.sales.budget.line')
        self.assertTrue(any(v[1] == 'graph' for v in action['views']))

    def test_03_month_close_note_lists_top_gaps(self):
        self.team.user_id = self.env.user.id
        # Presupuesto grande, facturación baja → brecha; PA mayor brecha que PB.
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.pa.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 1000.0})
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.pb.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 200.0})
        budget.state = 'aprobado'
        self._invoice(self.pa, date(2040, 6, 10), 10.0, 10.0)  # 100 facturado
        budget.action_refresh_actuals()
        self.Cron._sgi_sales_budget_month_close(date(2040, 6, 1), date(2040, 6, 30))
        # El cierre de mes (P-A28 5.5) agenda dos avisos: la justificación del
        # incumplimiento y el de la brecha por producto. Tomamos el de la brecha.
        acts = self.env['mail.activity'].search([
            ('res_model', '=', 'sgi.sales.budget'), ('res_id', '=', budget.id)])
        top = acts.filtered(lambda a: 'mayor brecha' in (a.note or ''))
        self.assertTrue(top, "Debe existir el aviso con los productos de mayor brecha.")
        self.assertIn('PA-AN', top.note, "El producto con mayor brecha aparece.")


@tagged('post_install', '-at_install')
class TestSalesBudgetAnalysisViews(TransactionCase):
    """5.3: submenú Análisis con 4 vistas; anti-doble conteo; botón de ficha."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado analisis 53'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Prod 53', 'type': 'consu', 'uom_id': cls.uom_m.id,
            'list_price': 5.0})

    def _budget_line(self, year, state):
        budget = self.Budget.create({'year': year, 'team_id': self.team.id})
        line = self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(year, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        if state != 'borrador':
            budget.state = state
        return budget, line

    def test_01_four_actions_and_views_validate(self):
        Line = self.env['sgi.sales.budget.line']
        for xmlid in ('sgi_sales_analysis_mercado_action',
                      'sgi_sales_analysis_cliente_action',
                      'sgi_sales_analysis_producto_action',
                      'sgi_sales_analysis_global_action'):
            action = self.env.ref('quimibond_sgi.%s' % xmlid)
            self.assertEqual(action.res_model, 'sgi.sales.budget.line')
            self.assertTrue(action.view_ids, "La acción define sus vistas.")
            for v in action.view_ids:
                # get_view valida/renderiza la arquitectura de cada vista.
                Line.get_view(view_id=v.view_id.id, view_type=v.view_mode)
            ctx = action.context
            self.assertIn('search_default_vigente', ctx)
            self.assertIn('search_default_presupuesto', ctx)

    def test_02_vigente_filter_excludes_obsolete_and_draft(self):
        appr, appr_line = self._budget_line(2040, 'aprobado')
        obso, obso_line = self._budget_line(2041, 'obsoleto')
        draft, draft_line = self._budget_line(2042, 'borrador')
        vigente = self.Line.search([
            ('budget_id.state', '=', 'aprobado'),
            ('team_id', '=', self.team.id)])
        self.assertIn(appr_line, vigente)
        self.assertNotIn(obso_line, vigente, "Obsoleto excluido por defecto.")
        self.assertNotIn(draft_line, vigente, "Borrador excluido por defecto.")

    def test_03_ficha_button_filters_by_budget(self):
        appr, appr_line = self._budget_line(2040, 'aprobado')
        other, other_line = self._budget_line(2041, 'aprobado')
        action = appr.action_open_analysis()
        self.assertEqual(action['res_model'], 'sgi.sales.budget.line')
        self.assertIn(('budget_id', '=', appr.id), action['domain'])
        self.assertNotIn('search_default_vigente', action.get('context', {}))
        lines = self.Line.search(action['domain'])
        self.assertIn(appr_line, lines)
        self.assertNotIn(other_line, lines)


@tagged('post_install', '-at_install')
class TestSalesBudgetPriceControl(TransactionCase):
    """5.4: control de precios (lista vs facturado) y cobertura de listas."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Param = cls.env['ir.config_parameter'].sudo()
        cls.team = cls.env['crm.team'].create({'name': 'Mercado precio 54'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.usd = cls.env.ref('base.USD')  # moneda de la compañía
        cls.mxn = cls.env.ref('base.MXN')
        cls.mxn.active = True
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        cls.product = cls.env['product.product'].create({
            'name': 'Tela precio 54', 'type': 'consu', 'uom_id': cls.uom_m.id,
            'list_price': 100.0})

    def _pricelist(self, currency, price):
        pl = self.env['product.pricelist'].create(
            {'name': 'PL %s 54' % currency.name, 'currency_id': currency.id})
        self.env['product.pricelist.item'].create({
            'pricelist_id': pl.id, 'applied_on': '1_product',
            'product_tmpl_id': self.product.product_tmpl_id.id,
            'compute_price': 'fixed', 'fixed_price': price})
        return pl

    def _client(self, pricelist):
        p = self.env['res.partner'].create(
            {'name': 'Cliente 54', 'is_company': True})
        p.property_product_pricelist = pricelist
        return p

    def _invoice(self, client, qty, price, currency, when=None):
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': client.id,
            'invoice_date': when or date(2040, 6, 10), 'team_id': self.team.id,
            'currency_id': currency.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': self.product.id, 'quantity': qty,
                'product_uom_id': self.uom_m.id, 'price_unit': price,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    def _line(self, budget, client):
        return self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id,
            'partner_id': client.id, 'qty_budget': 10.0})

    def test_01_gap_usd_list_usd_invoice(self):
        # Lista USD 100; factura USD a 110 → gap +10% en USD, sin cruce de divisas.
        client = self._client(self._pricelist(self.usd, 100.0))
        self._invoice(client, 5.0, 110.0, self.usd)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        line = self._line(budget, client)
        self.assertAlmostEqual(line.price_unit_currency, 100.0, places=2)
        self.assertAlmostEqual(line.price_real_unit_currency, 110.0, places=2)
        self.assertAlmostEqual(line.price_gap_pct, 10.0, places=2)
        self.assertFalse(line.price_gap_fx)

    def test_02_gap_crosses_fx_with_planning_rate(self):
        # Lista MXN (≠ USD compañía) 2000; factura en USD (compañía). El real
        # contable (USD) se convierte a MXN con el tipo presupuestal 20.
        self.Param.set_param('quimibond_sgi.budget_planning_rate', '20')
        client = self._client(self._pricelist(self.mxn, 2000.0))
        # 5 m a 110 USD c/u (moneda compañía) → 110 × 20 = 2200 MXN por unidad.
        self._invoice(client, 5.0, 110.0, self.usd)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        line = self._line(budget, client)
        self.assertEqual(line.list_currency_id, self.mxn)
        self.assertTrue(line.price_gap_fx, "Factura en otra moneda: cruza divisas.")
        self.assertAlmostEqual(line.price_real_unit_currency, 2200.0, places=1)
        self.assertAlmostEqual(line.price_gap_pct, 10.0, places=1)

    def test_03_thresholds_from_settings(self):
        client = self._client(self._pricelist(self.usd, 100.0))
        self._invoice(client, 5.0, 105.0, self.usd)  # +5%
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        line = self._line(budget, client)
        # Default tol 3 / grave 10 → 5% = leve.
        self.assertEqual(line.price_gap_alert, 'leve')
        # Sube la tolerancia a 8 → 5% pasa a OK (tras refresh).
        self.Param.set_param('quimibond_sgi.price_gap_tolerance_pct', '8.0')
        budget.action_refresh_actuals()
        self.assertEqual(line.price_gap_alert, 'ok')
        # Baja el grave a 4 → 5% pasa a grave.
        self.Param.set_param('quimibond_sgi.price_gap_tolerance_pct', '3.0')
        self.Param.set_param('quimibond_sgi.price_gap_grave_pct', '4.0')
        budget.action_refresh_actuals()
        self.assertEqual(line.price_gap_alert, 'grave')

    def test_04_approve_blocked_without_price_coverage(self):
        # Producto sin precio en la lista aplicable → no_price_count > 0.
        prod0 = self.env['product.product'].create(
            {'name': 'Sin precio 54', 'type': 'consu', 'uom_id': self.uom_m.id,
             'list_price': 0.0})
        client = self.env['res.partner'].create(
            {'name': 'Cli sin precio', 'is_company': True})
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': prod0.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id,
            'partner_id': client.id, 'qty_budget': 5.0})
        self.assertTrue(budget.no_price_count > 0)
        # El Admin de ventas manda a revisión: aquí el gate solo AVISA (no bloquea).
        sale_mgr = self.env['res.users'].create({
            'name': 'Vendedor 54', 'login': 'vend54',
            'group_ids': [(6, 0, [self.env.ref('sales_team.group_sale_manager').id])]})
        budget.with_user(sale_mgr).action_send_to_review()
        self.assertEqual(budget.state, 'revisado')
        # La alta dirección aprueba: aquí el gate BLOQUEA sin excepción.
        director = self.env['res.users'].create({
            'name': 'Dirección 54', 'login': 'dir54',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_director').id])]})
        with self.assertRaises(UserError):
            budget.with_user(director).action_approve()
        # Con contexto de excepción → aprueba y deja constancia.
        msgs_before = len(budget.message_ids)
        budget.with_user(director).with_context(
            sgi_bypass_price_check=True).action_approve()
        self.assertEqual(budget.state, 'aprobado')
        self.assertTrue(len(budget.message_ids) > msgs_before,
                        "El bypass deja constancia en el chatter.")

    def test_04b_manager_approves_with_exception(self):
        prod0 = self.env['product.product'].create(
            {'name': 'Sin precio 54b', 'type': 'consu', 'uom_id': self.uom_m.id,
             'list_price': 0.0})
        client = self.env['res.partner'].create(
            {'name': 'Cli sp b', 'is_company': True})
        manager = self.env['res.users'].create({
            'name': 'MAST 54', 'login': 'mast54',
            'group_ids': [(6, 0, [self.env.ref('quimibond_sgi.group_sgi_manager').id])]})
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': prod0.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id,
            'partner_id': client.id, 'qty_budget': 5.0})
        budget.with_user(manager).action_send_to_review()
        # El Jefe MAST/SGI (fallback de sistema) aprueba con excepción explícita.
        budget.with_user(manager).with_context(
            sgi_bypass_price_check=True).action_approve()
        self.assertEqual(budget.state, 'aprobado')

    def test_05_refresh_does_not_change_frozen_price(self):
        client = self._client(self._pricelist(self.usd, 100.0))
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        line = self._line(budget, client)
        self.assertAlmostEqual(line.price_unit_currency, 100.0, places=2)
        budget.state = 'aprobado'
        # Cambia la lista y refresca: el precio congelado NO se toca (la referencia).
        client.property_product_pricelist.item_ids.fixed_price = 200.0
        self._invoice(client, 5.0, 130.0, self.usd)
        budget.action_refresh_actuals()
        self.assertAlmostEqual(line.price_unit_currency, 100.0, places=2,
                               msg="Precio de lista congelado en aprobado.")
        # Pero el real y el gap sí se actualizan.
        self.assertAlmostEqual(line.price_real_unit_currency, 130.0, places=2)

    def test_06_coverage_report_and_deviation_count(self):
        client = self._client(self._pricelist(self.usd, 100.0))
        self._invoice(client, 5.0, 130.0, self.usd)  # +30% → grave
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        self._line(budget, client)
        self.assertEqual(budget.price_deviation_count, 1)
        act = budget.action_view_price_deviations()
        self.assertEqual(act['res_model'], 'sgi.sales.budget.line')
        cov = budget.action_price_coverage_report()
        self.assertIn(('has_list_price', '=', False), cov['domain'])


@tagged('post_install', '-at_install')
class TestSalesBudgetLifecycle55(TransactionCase):
    """Mini-fase 5.5 — alineación al P-A28 Rev.15: ciclos, roles e interconexión.
    Presupuesto: borrador→revisado→aprobado (alta dirección). Pronóstico:
    documento vivo borrador↔revisado (sin candado). Precarga combinada, MPS sin
    doble conteo y revaluación de junio (Nota 1)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Cron = cls.env['sgi.cron']
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        cls.product = cls.env['product.product'].create({
            'name': 'Prod 55 P', 'default_code': 'P55', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 10.0})
        cls.product_q = cls.env['product.product'].create({
            'name': 'Prod 55 Q', 'default_code': 'Q55', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 5.0})
        _seed_budget_pricelist(cls.env, 10.0)  # líneas globales a 10 (P5)
        cls.client = cls.env['res.partner'].create(
            {'name': 'Cliente 55', 'is_company': True})
        cls.sale_mgr = cls.env['res.users'].create({
            'name': 'Admin ventas 55', 'login': 'ventas55',
            'group_ids': [(6, 0, [cls.env.ref('sales_team.group_sale_manager').id])]})
        cls.director = cls.env['res.users'].create({
            'name': 'Dirección 55', 'login': 'dir55',
            'group_ids': [(6, 0, [cls.env.ref('quimibond_sgi.group_sgi_director').id])]})
        cls.raso = cls.env['res.users'].create({
            'name': 'Raso 55', 'login': 'raso55',
            'group_ids': [(6, 0, [cls.env.ref('quimibond_sgi.group_sgi_user').id])]})

    def _team(self, name):
        return self.env['crm.team'].create({'name': name})

    def _monday(self, ref):
        from datetime import timedelta
        return ref - timedelta(days=ref.weekday())

    def _forecast(self, team, partner=None):
        return self.Budget.create({
            'year': 2040, 'team_id': team.id, 'kind': 'pronostico',
            'partner_id': (partner or self.client).id})

    def _invoice(self, team, partner, product, when, qty):
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': partner.id,
            'invoice_date': when, 'team_id': team.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': product.id, 'quantity': qty,
                'product_uom_id': self.uom_m.id, 'price_unit': 10.0,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    # (i) write en línea de un pronóstico revisado → vuelve a borrador y postea.
    def test_01_forecast_line_write_reopens_and_posts(self):
        team = self._team('L55 reopen')
        fc = self._forecast(team)
        monday = self._monday(date(2040, 6, 3))
        line = self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id, 'date': monday,
            'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        fc.with_user(self.sale_mgr).action_send_to_review()
        self.assertEqual(fc.state, 'revisado')
        before = len(fc.message_ids)
        line.write({'qty_budget': 20.0})
        self.assertEqual(fc.state, 'borrador', "Editar líneas reabre el pronóstico.")
        self.assertTrue(len(fc.message_ids) > before)
        self.assertTrue(any('requiere revisión' in (m.body or '')
                            for m in fc.message_ids))

    # (ii) el pronóstico no admite action_revise.
    def test_02_forecast_has_no_revise(self):
        team = self._team('L55 no revise')
        fc = self._forecast(team)
        self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id,
            'date': self._monday(date(2040, 6, 3)), 'uom_id': self.uom_m.id,
            'qty_budget': 10.0})
        with self.assertRaises(UserError):
            fc.action_revise()

    # (iii) presupuesto: requiere revisado antes de aprobado y grupo Dirección.
    def test_03_budget_review_then_director_approves(self):
        team = self._team('L55 flow')
        budget = self.Budget.create({'year': 2040, 'team_id': team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        # Desde borrador no se aprueba: falta el paso 'revisado'.
        with self.assertRaises(UserError):
            budget.with_user(self.director).action_approve()
        budget.with_user(self.sale_mgr).action_send_to_review()
        self.assertEqual(budget.state, 'revisado')
        # El Admin de ventas revisa pero NO aprueba (es de la alta dirección).
        with self.assertRaises(UserError):
            budget.with_user(self.sale_mgr).action_approve()
        with self.assertRaises(UserError):
            budget.with_user(self.raso).action_approve()
        budget.with_user(self.director).action_approve()
        self.assertEqual(budget.state, 'aprobado')

    # (iv) precarga combina pronóstico (semana→mes) + facturado real (12 m).
    def test_04_preload_combines_forecast_and_real(self):
        from datetime import timedelta
        team = self._team('L55 preload')
        fc = self._forecast(team)
        # El real cae en el mes ACTUAL (dentro de los 12 meses); el pronóstico se
        # coloca en el MISMO mes de 2040 para que ambos aporten a la misma celda.
        today = fields.Date.context_today(self.env['sgi.cron'])
        when = today.replace(day=15)
        first = date(2040, when.month, 1)
        first_monday = first + timedelta(days=(7 - first.weekday()) % 7)
        # Pronóstico: 2 semanas × 15 m = 30 m para el producto P en ese mes.
        for monday in (first_monday, first_monday + timedelta(days=7)):
            self.Line.create({
                'budget_id': fc.id, 'product_id': self.product.id,
                'date': monday, 'uom_id': self.uom_m.id, 'qty_budget': 15.0})
        fc.state = 'revisado'
        self._invoice(team, self.client, self.product, when, 20.0)
        self._invoice(team, self.client, self.product_q, when, 25.0)
        budget = self.Budget.create({'year': 2040, 'team_id': team.id})
        budget.action_preload_from_orders()
        month = when.month
        line_p = budget.line_ids.filtered(lambda l: l.product_id == self.product)
        line_q = budget.line_ids.filtered(lambda l: l.product_id == self.product_q)
        self.assertEqual(len(line_p), 1)
        self.assertEqual(line_p.date, date(2040, month, 1))
        self.assertEqual(line_p.partner_id, self.client)
        self.assertEqual(line_p.qty_budget, 30.0,
                         "max(pronóstico 30, real 20) = 30.")
        self.assertEqual(line_q.qty_budget, 25.0,
                         "Sin pronóstico: manda el real 25.")

    # (v) send_to_mps del presupuesto omite productos con pronóstico vigente.
    def test_05_budget_mps_omits_forecast_covered(self):
        team = self._team('L55 mps')
        fc = self._forecast(team)
        self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id,
            'date': self._monday(date(2040, 6, 3)), 'uom_id': self.uom_m.id,
            'qty_budget': 10.0})
        fc.state = 'revisado'
        budget = self.Budget.create({'year': 2040, 'team_id': team.id})
        for prod in (self.product, self.product_q):
            self.Line.create({
                'budget_id': budget.id, 'product_id': prod.id,
                'date': date(2040, 6, 1), 'uom_id': self.uom_m.id,
                'qty_budget': 100.0})
        covered = budget._sgi_forecast_covered_products()
        self.assertIn(self.product.id, covered, "P lo cubre el pronóstico.")
        self.assertNotIn(self.product_q.id, covered, "Q no.")
        if 'mrp.production.schedule' not in self.env:
            self.skipTest("mrp_mps no está instalado en esta BD.")
        budget.state = 'aprobado'
        before = len(budget.message_ids)
        budget.action_send_to_mps()
        self.assertTrue(len(budget.message_ids) > before)
        bodies = " ".join(m.body or '' for m in budget.message_ids)
        self.assertIn('omitidos 1', bodies)
        self.assertIn('P55', bodies, "El producto omitido aparece en el resumen.")

    # (vi) cron de junio crea la actividad de revaluación al Admin de ventas.
    def test_06_june_revaluation_creates_activity(self):
        team = self._team('L55 junio')
        budget = self.Budget.create({'year': 2040, 'team_id': team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2040, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        budget.state = 'aprobado'
        self.Cron._sgi_sales_budget_revaluation(2040)
        acts = self.env['mail.activity'].search([
            ('res_model', '=', 'sgi.sales.budget'),
            ('res_id', '=', budget.id),
            ('summary', '=', "Revaluar cantidades del S2 — P-A28 Nota 1")])
        self.assertTrue(acts, "Debe agendar la revaluación del S2 (P-A28 Nota 1).")

    # (vii) sin candado en pronóstico: editar qty no levanta el UserError del lock.
    def test_07_forecast_no_lock_on_qty_edit(self):
        team = self._team('L55 nolock')
        fc = self._forecast(team)
        line = self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id,
            'date': self._monday(date(2040, 6, 3)), 'uom_id': self.uom_m.id,
            'qty_budget': 10.0})
        fc.state = 'revisado'
        # Un Admin de ventas (NO Jefe MAST) edita la cantidad: el candado le
        # aplicaría si la línea estuviera bloqueada, pero el pronóstico es
        # documento vivo → sin candado, no levanta UserError.
        line.with_user(self.sale_mgr).write({'qty_budget': 50.0})
        self.assertEqual(line.qty_budget, 50.0)
        self.assertEqual(fc.state, 'borrador', "La edición reabrió el pronóstico.")


@tagged('post_install', '-at_install')
class TestSalesBudgetImportPartner(TransactionCase):
    """Hotfix: resolución de cliente ambigua en el import mensual."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Wizard = cls.env['sgi.sales.budget.import']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado import cli'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.prod = cls.env['product.product'].create({
            'name': 'Prod cli', 'default_code': 'HFXPROD', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.fxi = cls.env['res.partner'].create(
            {'name': 'FXITEST', 'is_company': True})
        cls.fxi_cua = cls.env['res.partner'].create(
            {'name': 'FXITEST CUAUTITLAN', 'is_company': True})

    def _import(self, rows):
        import base64
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Presupuesto'
        ws.append(['PRODUCTO', 'CLIENTE', 'ENERO m'])
        for code, client, qty in rows:
            ws.append([code, client, qty])
        buf = io.BytesIO()
        wb.save(buf)
        budget = self.Budget.create({'year': 2040, 'team_id': self.team.id})
        wiz = self.Wizard.create({
            'budget_id': budget.id, 'file': base64.b64encode(buf.getvalue()),
            'sheet_name': 'Presupuesto'})
        wiz.action_import()
        return budget, wiz

    def test_01_exact_wins_over_prefix(self):
        budget, wiz = self._import([
            ('HFXPROD', 'FXITEST', 100),
            ('HFXPROD', 'FXITEST CUAUTITLAN', 200)])
        by_partner = {l.partner_id: l.qty_budget for l in budget.line_ids}
        self.assertEqual(by_partner.get(self.fxi), 100.0,
                         "'FXITEST' resuelve al exacto, no al que lo contiene.")
        self.assertEqual(by_partner.get(self.fxi_cua), 200.0)

    def test_02_unique_ilike_resolves(self):
        uniq = self.env['res.partner'].create(
            {'name': 'ZZUNIQUECLI COMPANY', 'is_company': True})
        budget, wiz = self._import([('HFXPROD', 'ZZUNIQUECLI', 50)])
        line = budget.line_ids
        self.assertEqual(len(line), 1)
        self.assertEqual(line.partner_id, uniq, "ilike único sí resuelve.")

    def test_03_ambiguous_ilike_skips_and_reports(self):
        self.env['res.partner'].create({'name': 'ZAMBI NORTE', 'is_company': True})
        self.env['res.partner'].create({'name': 'ZAMBI SUR', 'is_company': True})
        budget, wiz = self._import([('HFXPROD', 'ZAMBI', 300)])
        self.assertFalse(budget.line_ids, "Cliente ambiguo: no se importa la fila.")
        self.assertIn('ambiguo', wiz.result)
        self.assertIn('ZAMBI', wiz.result)

    def test_04_typographic_apostrophe_resolves(self):
        osu = self.env['res.partner'].create(
            {'name': "O'SULLIVAN", 'is_company': True})
        # El template trae el apóstrofe tipográfico ’.
        budget, wiz = self._import([('HFXPROD', 'O’SULLIVAN', 75)])
        line = budget.line_ids
        self.assertEqual(len(line), 1)
        self.assertEqual(line.partner_id, osu)
        self.assertEqual(line.qty_budget, 75.0)

    def test_05_client_not_found_skips(self):
        budget, wiz = self._import([('HFXPROD', 'CLIENTE INEXISTENTE ZZZ', 10)])
        self.assertFalse(budget.line_ids, "Cliente no encontrado: fila reportada.")
        self.assertIn('no encontrado', wiz.result)


@tagged('post_install', '-at_install')
class TestSalesBudgetForecastCoverage(TransactionCase):
    """5.6: cobertura del pronóstico (pedidos vs pronosticado, P-A28 4.2.2.7)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from datetime import timedelta
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Cron = cls.env['sgi.cron']
        cls.Param = cls.env['ir.config_parameter'].sudo()
        cls.team = cls.env['crm.team'].create({'name': 'Mercado cobertura'})
        cls.team.user_id = cls.env.user.id
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.client = cls.env['res.partner'].create(
            {'name': 'Cliente cobertura', 'is_company': True})
        today = date.today()
        cls.cur_mon = today - timedelta(days=today.weekday())

    def _monday(self, weeks):
        from datetime import timedelta
        return self.cur_mon + timedelta(weeks=weeks)

    def _product(self, code):
        return self.env['product.product'].create({
            'name': 'P %s' % code, 'default_code': code, 'type': 'consu',
            'uom_id': self.uom_m.id, 'list_price': 10.0})

    def _order(self, product, qty, monday):
        from datetime import datetime, timedelta
        so = self.env['sale.order'].create({
            'partner_id': self.client.id, 'team_id': self.team.id,
            'commitment_date': datetime.combine(
                monday + timedelta(days=1), datetime.min.time()),
            'order_line': [(0, 0, {
                'product_id': product.id, 'product_uom_qty': qty,
                'product_uom_id': self.uom_m.id, 'price_unit': 1.0})]})
        so.action_confirm()
        return so

    def _forecast(self):
        return self.Budget.create({
            'year': self.cur_mon.year, 'team_id': self.team.id,
            'kind': 'pronostico', 'partner_id': self.client.id})

    def _fline(self, budget, product, monday, qty):
        return self.Line.create({
            'budget_id': budget.id, 'product_id': product.id, 'date': monday,
            'uom_id': self.uom_m.id, 'qty_budget': qty})

    def test_01_coverage_states(self):
        p1, p2, p3, p4, p5 = [self._product('COV%d' % i) for i in range(1, 6)]
        self._order(p1, 100, self._monday(0))   # cubierto
        self._order(p2, 60, self._monday(0))    # parcial
        self._order(p4, 120, self._monday(0))   # excedido
        self._order(p5, 100, self._monday(3))   # fuera de horizonte (>= horizonte 3)
        fc = self._forecast()
        l1 = self._fline(fc, p1, self._monday(0), 100)
        l2 = self._fline(fc, p2, self._monday(0), 100)
        l3 = self._fline(fc, p3, self._monday(0), 100)  # sin pedido
        l4 = self._fline(fc, p4, self._monday(0), 100)
        l5 = self._fline(fc, p5, self._monday(3), 100)
        self.assertEqual(l1.coverage_state, 'cubierto')
        self.assertEqual(l2.coverage_state, 'parcial')
        self.assertEqual(l3.coverage_state, 'sin_pedido')
        self.assertEqual(l4.coverage_state, 'excedido')
        self.assertEqual(l5.coverage_state, 'fuera_horizonte')
        self.assertEqual(fc.uncovered_count, 2)  # l2 parcial + l3 sin_pedido

    def test_02_cron_one_activity_idempotent(self):
        p1 = self._product('CR1')
        fc = self._forecast()
        self._fline(fc, p1, self._monday(0), 100)  # sin pedido → descubierta
        self.Cron.cron_forecast_coverage()
        acts = self.env['mail.activity'].search([
            ('res_model', '=', 'sgi.sales.budget'), ('res_id', '=', fc.id)])
        self.assertEqual(len(acts), 1, "Una sola actividad por pronóstico.")
        self.assertIn('CR1', acts.note or '')
        self.assertIn('faltante', acts.note or '')
        self.Cron.cron_forecast_coverage()  # idempotente
        self.assertEqual(self.env['mail.activity'].search_count([
            ('res_model', '=', 'sgi.sales.budget'), ('res_id', '=', fc.id)]), 1)

    def test_03_create_draft_quotation(self):
        p1 = self._product('QT1')
        self._order(p1, 40, self._monday(0))  # parcial: faltan 60
        fc = self._forecast()
        line = self._fline(fc, p1, self._monday(0), 100)
        self.assertEqual(line.coverage_state, 'parcial')
        action = line.action_create_draft_quotation()
        order = self.env['sale.order'].browse(action['res_id'])
        self.assertEqual(order.state, 'draft', "NO se confirma la cotización.")
        self.assertEqual(order.origin, fc.folio)
        self.assertEqual(order.order_line.product_uom_qty, 60.0, "Solo el faltante.")
        self.assertEqual(order.commitment_date.date(), self._monday(0))
        # Reabrir en vez de duplicar.
        action2 = line.action_create_draft_quotation()
        self.assertEqual(action2['res_id'], order.id)

    def test_04_orders_without_forecast(self):
        p1, p2 = self._product('OWF1'), self._product('OWF2')
        self._order(p1, 50, self._monday(0))  # con línea de pronóstico
        self._order(p2, 30, self._monday(0))  # SIN línea de pronóstico
        fc = self._forecast()
        self._fline(fc, p1, self._monday(0), 100)
        orphans = fc._sgi_orders_without_forecast()
        self.assertIn(p2, orphans.mapped('product_id'))
        self.assertNotIn(p1, orphans.mapped('product_id'))

    def test_05_horizon_and_tolerance_params(self):
        p1 = self._product('HZ1')
        self._order(p1, 105, self._monday(2))  # 1.05
        fc = self._forecast()
        line = self._fline(fc, p1, self._monday(2), 100)
        # Horizonte default 3 → semana +2 está dentro; 1.05 <= 1.10 → cubierto.
        self.assertEqual(line.coverage_state, 'cubierto')
        # Tolerancia a 3% → 1.05 > 1.03 → excedido (tras refresh).
        self.Param.set_param('quimibond_sgi.forecast_over_tolerance_pct', '3.0')
        fc.action_refresh_actuals()
        self.assertEqual(line.coverage_state, 'excedido')
        # Horizonte a 1 → semana +2 queda fuera.
        self.Param.set_param('quimibond_sgi.forecast_capture_horizon_weeks', '1')
        fc.action_refresh_actuals()
        self.assertEqual(line.coverage_state, 'fuera_horizonte')

    def test_06_no_coverage_for_presupuesto(self):
        p1 = self._product('PB1')
        budget = self.Budget.create({
            'year': self.cur_mon.year, 'team_id': self.team.id})
        line = self.Line.create({
            'budget_id': budget.id, 'product_id': p1.id,
            'date': date(self.cur_mon.year, 6, 1), 'uom_id': self.uom_m.id,
            'qty_budget': 100.0})
        self.assertEqual(line.coverage_state, 'fuera_horizonte')
        self.assertEqual(budget.uncovered_count, 0)


@tagged('post_install', '-at_install')
class TestSalesBudgetPricePlausibility(TransactionCase):
    """Ajuste 13.10 P1: detección REAL de 'sin precio de lista' (regla + umbral)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Param = cls.env['ir.config_parameter'].sudo()
        cls.team = cls.env['crm.team'].create({'name': 'Mercado plausible'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.usd = cls.env.ref('base.USD')  # moneda compañía en esta BD
        cls.mxn = cls.env.ref('base.MXN')
        cls.mxn.active = True
        cls.product = cls.env['product.product'].create({
            'name': 'Tela plausible', 'default_code': 'PLAUS', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 1.0})  # placeholder $1

    def _client(self, pricelist):
        p = self.env['res.partner'].create(
            {'name': 'Cli plaus', 'is_company': True})
        p.property_product_pricelist = pricelist
        return p

    def _list(self, currency=None, rules=True, fixed=None, applied='1_product',
              base='fixed'):
        pl = self.env['product.pricelist'].create({
            'name': 'PL plaus', 'currency_id': (currency or self.usd).id})
        if rules:
            vals = {'pricelist_id': pl.id, 'applied_on': applied,
                    'compute_price': base}
            if applied == '1_product':
                vals['product_tmpl_id'] = self.product.product_tmpl_id.id
            if base == 'fixed':
                vals['fixed_price'] = fixed
            self.env['product.pricelist.item'].create(vals)
        return pl

    def _line(self, pricelist):
        budget = self.Budget.create({'year': 2044, 'team_id': self.team.id})
        return self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2044, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0,
            'partner_id': self._client(pricelist).id})

    def test_01_no_rule_falls_to_sale_price_not_valid(self):
        # Lista SIN regla para el producto: el engine cae al precio de venta ($1);
        # NO cuenta como precio de lista.
        line = self._line(self._list(rules=False))
        self.assertFalse(line.has_list_price)
        self.assertIn('SIN REGLA', line.price_source)

    def test_02_real_rule_normal_price_valid(self):
        line = self._line(self._list(fixed=100.0))
        self.assertTrue(line.has_list_price)
        self.assertIn("Lista", line.price_source)

    def test_03_rule_below_min_plausible_is_placebo(self):
        # Regla real pero a $0.50 (< umbral 5) → placebo, sin precio de lista.
        line = self._line(self._list(fixed=0.50))
        self.assertFalse(line.has_list_price)
        self.assertIn('implausible', line.price_source)

    def test_04_usd_list_no_rule_the_0057_case(self):
        # Lista en USD sin regla, producto $1: 1/17.52 ≈ 0.057 (el caso real). El
        # $1 lavado por la conversión NO es precio de lista.
        line = self._line(self._list(currency=self.usd, rules=False))
        self.assertFalse(line.has_list_price)

    def test_05_min_plausible_configurable(self):
        # Sube el umbral a 150 → una regla a 100 pasa a implausible.
        self.Param.set_param('quimibond_sgi.price_min_plausible', '150')
        line = self._line(self._list(fixed=100.0))
        self.assertFalse(line.has_list_price)


@tagged('post_install', '-at_install')
class TestSalesBudgetBudgetPricelist(TransactionCase):
    """Ajuste 13.10 P5: lista presupuestal para líneas SIN cliente."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.Param = cls.env['ir.config_parameter'].sudo()
        cls.team = cls.env['crm.team'].create({'name': 'Mercado ppto PL'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Tela ppto PL', 'default_code': 'PPTOPL', 'type': 'consu',
            'uom_id': cls.uom_m.id, 'list_price': 7.0})

    def _global_line(self):
        budget = self.Budget.create({'year': 2045, 'team_id': self.team.id})
        return self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2045, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})

    def test_01_without_param_no_price(self):
        self.Param.set_param('quimibond_sgi.budget_pricelist_id', '0')
        line = self._global_line()
        self.assertFalse(line.has_list_price)
        self.assertIn('SIN LISTA PRESUPUESTAL', line.price_source)

    def test_02_with_param_uses_that_list(self):
        pl = self.env['product.pricelist'].create({'name': 'Presupuestal PPV'})
        self.env['product.pricelist.item'].create({
            'pricelist_id': pl.id, 'applied_on': '1_product',
            'product_tmpl_id': self.product.product_tmpl_id.id,
            'compute_price': 'fixed', 'fixed_price': 42.0})
        self.Param.set_param('quimibond_sgi.budget_pricelist_id', pl.id)
        line = self._global_line()
        self.assertTrue(line.has_list_price)
        self.assertAlmostEqual(line.price_unit_budget, 42.0, places=2)
        self.assertIn('Presupuestal PPV', line.price_source)

    def test_03_client_lines_unaffected(self):
        # Con o sin lista presupuestal, las líneas CON cliente usan la del cliente.
        self.Param.set_param('quimibond_sgi.budget_pricelist_id', '0')
        pl = self.env['product.pricelist'].create({'name': 'Cli PL PPV'})
        self.env['product.pricelist.item'].create({
            'pricelist_id': pl.id, 'applied_on': '1_product',
            'product_tmpl_id': self.product.product_tmpl_id.id,
            'compute_price': 'fixed', 'fixed_price': 55.0})
        client = self.env['res.partner'].create(
            {'name': 'Cli PPTO', 'is_company': True})
        client.property_product_pricelist = pl
        budget = self.Budget.create({'year': 2045, 'team_id': self.team.id})
        line = self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2045, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0,
            'partner_id': client.id})
        self.assertTrue(line.has_list_price)
        self.assertAlmostEqual(line.price_unit_budget, 55.0, places=2)


@tagged('post_install', '-at_install')
class TestSalesBudgetReviewedGovernance(TransactionCase):
    """Ajuste 13.10 P2: gobernanza del estado 'revisado' del presupuesto."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado gob'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Tela gob', 'default_code': 'GOB', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.sale_mgr = cls.env['res.users'].create({
            'name': 'Admin ventas gob', 'login': 'admgob',
            'group_ids': [(6, 0, [
                cls.env.ref('sales_team.group_sale_manager').id])]})

    def _reviewed_budget(self):
        budget = self.Budget.create({'year': 2046, 'team_id': self.team.id})
        line = self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2046, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        budget.with_user(self.sale_mgr).action_send_to_review()
        self.assertEqual(budget.state, 'revisado')
        return budget, line

    def test_01_write_reopens_to_borrador(self):
        budget, line = self._reviewed_budget()
        before = len(budget.message_ids)
        line.write({'qty_budget': 20.0})
        self.assertEqual(budget.state, 'borrador',
                         "Editar captura tras la revisión reabre el presupuesto.")
        self.assertTrue(len(budget.message_ids) > before)
        self.assertTrue(any('re-revisión' in (m.body or '')
                            for m in budget.message_ids))

    def test_02_create_line_reopens(self):
        budget, _line = self._reviewed_budget()
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.product.id,
            'date': date(2046, 7, 1), 'uom_id': self.uom_m.id, 'qty_budget': 5.0})
        self.assertEqual(budget.state, 'borrador')

    def test_03_unlink_line_reopens(self):
        budget, line = self._reviewed_budget()
        line.unlink()
        self.assertEqual(budget.state, 'borrador')

    def test_04_refresh_does_not_reopen(self):
        budget, _line = self._reviewed_budget()
        budget.action_refresh_actuals()
        self.assertEqual(budget.state, 'revisado',
                         "Refrescar la foto real NO cuenta como edición de captura.")


@tagged('post_install', '-at_install')
class TestSalesBudgetForecastCoveredDraft(TransactionCase):
    """Ajuste 13.10 P3: solo un pronóstico REVISADO omite productos del MPS."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado cov draft'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.product = cls.env['product.product'].create({
            'name': 'Tela cov draft', 'default_code': 'COVD', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.client = cls.env['res.partner'].create(
            {'name': 'Cli cov draft', 'is_company': True})

    def _forecast(self):
        from datetime import timedelta
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        fc = self.Budget.create({
            'year': monday.year, 'team_id': self.team.id, 'kind': 'pronostico',
            'partner_id': self.client.id})
        self.Line.create({
            'budget_id': fc.id, 'product_id': self.product.id, 'date': monday,
            'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        return fc

    def test_01_draft_forecast_does_not_cover(self):
        fc = self._forecast()  # queda en borrador
        budget = self.Budget.create({'year': fc.year, 'team_id': self.team.id})
        self.assertNotIn(self.product.id, budget._sgi_forecast_covered_products(),
                         "Un pronóstico en borrador NO omite el producto (no puede "
                         "enviar demanda; lo dejaría sin cubrir en el MPS).")
        # Al revisarlo, sí lo cubre.
        fc.state = 'revisado'
        self.assertIn(self.product.id, budget._sgi_forecast_covered_products())


@tagged('post_install', '-at_install')
class TestSalesBudgetReconcile(TransactionCase):
    """Ajuste 13.10 P4: conciliación facturado vs contabilidad (A = B + partidas)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.team = cls.env['crm.team'].create({'name': 'Mercado concilia'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income')], limit=1)
        cls.p_pres = cls.env['product.product'].create({
            'name': 'Prod presupuestado', 'default_code': 'RP', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.p_unbud = cls.env['product.product'].create({
            'name': 'Prod no presupuestado', 'default_code': 'RU', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.cli_a = cls.env['res.partner'].create(
            {'name': 'Cliente A concilia', 'is_company': True})
        cls.cli_b = cls.env['res.partner'].create(
            {'name': 'Cliente B concilia', 'is_company': True})

    def _invoice(self, partner, lines, team=True, when=None, freight=0.0):
        ilines = []
        for product, qty, price in lines:
            ilines.append((0, 0, {
                'product_id': product.id, 'quantity': qty,
                'product_uom_id': self.uom_m.id, 'price_unit': price,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]}))
        if freight:
            ilines.append((0, 0, {
                'name': 'Flete', 'quantity': 1, 'price_unit': freight,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]}))
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': partner.id,
            'invoice_date': when or date(2047, 6, 10),
            'team_id': self.team.id if team else False,
            'invoice_line_ids': ilines})
        move.action_post()
        return move

    def test_01_reconcile_balances_exactly(self):
        # Presupuesto: p_pres POR CLIENTE cli_a.
        budget = self.Budget.create({'year': 2047, 'team_id': self.team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.p_pres.id,
            'date': date(2047, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0,
            'partner_id': self.cli_a.id})
        # Facturas del equipo:
        #  · cli_a p_pres 10×100 = 1000  (capturado, B)
        #  · cli_b p_pres 5×100  = 500   (presupuestado por cliente → a otro cliente)
        #  · cli_a p_unbud 3×50  = 150   (producto no presupuestado)
        #  · cli_a flete 200             (línea sin producto)
        self._invoice(self.cli_a, [(self.p_pres, 10, 100)])
        self._invoice(self.cli_b, [(self.p_pres, 5, 100)])
        self._invoice(self.cli_a, [(self.p_unbud, 3, 50)], freight=200.0)
        # Factura SIN equipo (informativo, fuera de A).
        self._invoice(self.cli_a, [(self.p_pres, 2, 100)], team=False)
        budget.action_refresh_actuals()  # la foto real alimenta B
        d = budget._sgi_reconcile_data()
        self.assertAlmostEqual(d['A'], 1850.0, places=2)  # 1000+500+150+200
        self.assertAlmostEqual(d['B'], 1000.0, places=2)
        self.assertAlmostEqual(d['by_client_others'], 500.0, places=2)
        self.assertAlmostEqual(d['unbudgeted'], 150.0, places=2)
        self.assertAlmostEqual(d['no_product'], 200.0, places=2)
        self.assertAlmostEqual(d['residual'], 0.0, places=2,
                               msg="A = B + partidas, residuo 0.")
        self.assertAlmostEqual(d['no_team'], 200.0, places=2)

    def test_02_action_posts_and_notifies(self):
        budget = self.Budget.create({'year': 2047, 'team_id': self.team.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': self.p_pres.id,
            'date': date(2047, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0,
            'partner_id': self.cli_a.id})
        self._invoice(self.cli_a, [(self.p_pres, 10, 100)])
        before = len(budget.message_ids)
        action = budget.action_reconcile_invoiced()
        self.assertEqual(action['tag'], 'display_notification')
        self.assertTrue(len(budget.message_ids) > before,
                        "Publica la tabla de conciliación en el chatter.")

    def test_03_reconcile_only_presupuesto(self):
        fc = self.Budget.create({
            'year': 2047, 'team_id': self.team.id, 'kind': 'pronostico',
            'partner_id': self.cli_a.id})
        with self.assertRaises(UserError):
            fc.action_reconcile_invoiced()


@tagged('post_install', '-at_install')
class TestSalesBudgetMultiCompany(TransactionCase):
    """Hotfix 13.11: el presupuesto solo mide la facturación/pedidos de SU
    compañía, y no truena con productos mal configurados de otra compañía."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.Budget = cls.env['sgi.sales.budget']
        cls.Line = cls.env['sgi.sales.budget.line']
        cls.company_a = cls.env.company
        cls.company_b = cls.env['res.company'].create({'name': 'Compañía B SGI'})
        cls.team = cls.env['crm.team'].create({'name': 'Mercado multi'})
        cls.uom_m = cls.env.ref('uom.product_uom_meter')
        cls.income = cls.env['account.account'].search(
            [('account_type', '=', 'income'),
             ('company_ids', 'in', cls.company_a.id)], limit=1)
        cls.product = cls.env['product.product'].create({
            'name': 'Tela multi', 'default_code': 'MULTI', 'type': 'consu',
            'uom_id': cls.uom_m.id})
        cls.client = cls.env['res.partner'].create(
            {'name': 'Cliente multi', 'is_company': True})

    def _invoice_a(self, product=None):
        move = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': self.client.id,
            'invoice_date': date(2048, 6, 10), 'team_id': self.team.id,
            'company_id': self.company_a.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': (product or self.product).id, 'quantity': 10.0,
                'product_uom_id': self.uom_m.id, 'price_unit': 100.0,
                'account_id': self.income.id, 'tax_ids': [(6, 0, [])]})]})
        move.action_post()
        return move

    def test_01_cross_company_invoice_does_not_leak(self):
        # Factura de la compañía A (10 × 100 = 1000). Un presupuesto de la
        # compañía B, mismo equipo/producto/cliente, NO la ve: el real filtra por
        # compañía (equivale a "la factura de B no entra al presupuesto de A").
        self._invoice_a()
        budget_a = self.Budget.create({
            'year': 2048, 'team_id': self.team.id, 'company_id': self.company_a.id})
        self.Line.create({
            'budget_id': budget_a.id, 'product_id': self.product.id,
            'date': date(2048, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        budget_a.action_refresh_actuals()
        self.assertAlmostEqual(budget_a.line_ids.qty_real, 10.0, places=2)
        self.assertAlmostEqual(budget_a.line_ids.amount_real, 1000.0, places=2)

        budget_b = self.Budget.create({
            'year': 2048, 'team_id': self.team.id, 'company_id': self.company_b.id})
        self.Line.create({
            'budget_id': budget_b.id, 'product_id': self.product.id,
            'date': date(2048, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 10.0})
        budget_b.action_refresh_actuals()
        self.assertAlmostEqual(
            budget_b.line_ids.qty_real, 0.0, places=2,
            msg="La facturación de otra compañía no entra al presupuesto.")
        self.assertAlmostEqual(budget_b.line_ids.amount_real, 0.0, places=2)

    def test_02_foreign_product_does_not_break_company_a_user(self):
        # Producto de la compañía B facturado en A (mal configurado): un usuario
        # solo con A refresca/precarga/concilia sin AccessError, y la conciliación
        # lo cuenta como NO presupuestado.
        self._invoice_a()  # producto aún compartido al timbrar
        # El producto pasa a ser de la compañía B tras timbrarse en A.
        self.env.cr.execute(
            "UPDATE product_template SET company_id=%s WHERE id=%s",
            (self.company_b.id, self.product.product_tmpl_id.id))
        self.env.invalidate_all()
        # Presupuesto de A con OTRO producto (propio); el foráneo entra solo por la
        # agregación del equipo, nunca como línea propia.
        prod_a = self.env['product.product'].create({
            'name': 'Prod A multi', 'default_code': 'PA-MULTI', 'type': 'consu',
            'uom_id': self.uom_m.id})
        budget = self.Budget.create({
            'year': 2048, 'team_id': self.team.id, 'company_id': self.company_a.id})
        self.Line.create({
            'budget_id': budget.id, 'product_id': prod_a.id,
            'date': date(2048, 6, 1), 'uom_id': self.uom_m.id, 'qty_budget': 5.0,
            'partner_id': self.client.id})
        user_a = self.env['res.users'].create({
            'name': 'Solo A', 'login': 'soloa_multi',
            'company_id': self.company_a.id,
            'company_ids': [(6, 0, [self.company_a.id])],
            'group_ids': [(6, 0, [
                self.env.ref('sales_team.group_sale_manager').id,
                self.env.ref('quimibond_sgi.group_sgi_manager').id,
                self.env.ref('account.group_account_invoice').id])]})
        budget_u = budget.with_user(user_a)
        # (a) refresco no truena con el producto foráneo en la agregación del equipo.
        budget_u.action_refresh_actuals()
        # (b) conciliación: no truena y el foráneo cae como NO presupuestado.
        d = budget_u._sgi_reconcile_data()
        self.assertAlmostEqual(d['unbudgeted'], 1000.0, places=2,
                               msg="El producto foráneo facturado cae como no "
                                   "presupuestado, sin reventar por multicompañía.")
        # (c) precarga: no truena y OMITE el producto foráneo (no lo presupuesta).
        budget_u.action_preload_from_orders()
        self.assertFalse(
            budget.line_ids.filtered(lambda l: l.product_id == self.product),
            "La precarga omite el producto de otra compañía (mal configurado).")
