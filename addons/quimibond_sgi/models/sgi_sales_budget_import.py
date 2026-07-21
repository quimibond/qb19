# -*- coding: utf-8 -*-
"""Asistente de importación del F-P-A28-18 desde el Excel real.

Lee la matriz producto × meses (pares "<mes> m" / "<mes> $") con openpyxl y crea
las líneas del presupuesto. Tolerante a mayúsculas/acentos/espacios en los
encabezados. Todo-o-nada por hoja para errores ESTRUCTURALES; los productos sin
match NO son error: se reportan y la importación sigue.

Convive con la importación plana estándar (base_import) de la lista de líneas.
"""
import base64
import io
import unicodedata
from datetime import date

from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError

_MONTHS = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'setiembre': 9, 'octubre': 10,
    'noviembre': 11, 'diciembre': 12,
}
_QTY_TOKENS = {'m', 'cant', 'cantidad', 'unidades', 'cantidades'}
_AMOUNT_TOKENS = {'$', 'importe', 'pesos', 'importes', 'mxn'}


def _norm(value):
    """Normaliza un encabezado: sin acentos, minúsculas, sin espacios extra."""
    if value is None:
        return ''
    text = str(value).strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')


class SgiSalesBudgetImport(models.TransientModel):
    _name = 'sgi.sales.budget.import'
    _description = "Importar presupuesto desde Excel (F-P-A28-18)"

    budget_id = fields.Many2one(
        'sgi.sales.budget', string="Presupuesto", required=True,
        default=lambda self: self.env.context.get('active_id'))
    file = fields.Binary(string="Archivo .xlsx", required=True)
    filename = fields.Char(string="Nombre del archivo")
    available_sheets = fields.Char(string="Hojas del archivo", readonly=True)
    sheet_name = fields.Char(
        string="Hoja a importar",
        help="Vacío = la primera hoja del libro.")
    conflict_mode = fields.Selection([
        ('replace', "Reemplazar líneas existentes del producto"),
        ('add', "Sumar a las existentes"),
    ], string="Si el producto ya tiene líneas", default='replace', required=True)
    result = fields.Text(string="Resultado", readonly=True)

    @api.onchange('file')
    def _onchange_file(self):
        self.available_sheets = False
        if not self.file:
            return
        try:
            wb = self._load_workbook()
        except UserError:
            self.available_sheets = "No se pudo leer el archivo (¿es .xlsx?)."
            return
        self.available_sheets = ", ".join(wb.sheetnames)
        if not self.sheet_name and wb.sheetnames:
            self.sheet_name = wb.sheetnames[0]

    def _load_workbook(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError("openpyxl no está disponible en este servidor.")
        try:
            # read_only=False: se necesita acceso aleatorio ws.cell(fila, col).
            return openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(self.file)), data_only=True)
        except Exception as exc:
            raise UserError("No se pudo abrir el Excel: %s" % exc)

    # --- Parseo del layout ---------------------------------------------------
    def _parse_header(self, ws):
        """Ubica la fila de encabezados (busca 'PRODUCTO') y mapea columnas."""
        header_row = prod_col = None
        max_r = min(ws.max_row or 0, 30)
        max_c = min(ws.max_column or 0, 80)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                if 'producto' in _norm(ws.cell(r, c).value):
                    header_row, prod_col = r, c
                    break
            if header_row:
                break
        if not header_row:
            raise UserError(
                "No se encontró la fila de encabezados (debe haber una celda "
                "'PRODUCTO'). Revisa la hoja seleccionada.")
        unit_col = client_col = None
        month_cols = {}
        for c in range(1, (ws.max_column or 0) + 1):
            if c == prod_col:
                continue
            text = _norm(ws.cell(header_row, c).value)
            if not text:
                continue
            if text == 'unidad':
                unit_col = c
                continue
            if text == 'cliente':
                client_col = c
                continue
            tokens = text.replace('$', ' $ ').split()
            month = next((_MONTHS[t] for t in tokens if t in _MONTHS), None)
            if not month:
                continue
            if _AMOUNT_TOKENS & set(tokens):
                month_cols[(month, 'amount')] = c
            elif _QTY_TOKENS & set(tokens):
                month_cols[(month, 'qty')] = c
        if not month_cols:
            raise UserError(
                "No se reconoció ninguna columna de mes (pares '<mes> m' / "
                "'<mes> $'). Revisa los encabezados.")
        return header_row, prod_col, unit_col, client_col, month_cols

    def _match_product(self, ref):
        ref = (ref or '').strip()
        if not ref:
            return self.env['product.product']
        Product = self.env['product.product']
        product = Product.search([('default_code', '=', ref)], limit=1)
        if product:
            return product
        product = Product.search([('name', '=', ref)], limit=1)
        if product:
            return product
        product = Product.search([('name', 'ilike', ref)], limit=2)
        return product if len(product) == 1 else Product

    @staticmethod
    def _as_number(value):
        """Número tolerante a comas de miles ('16,000' → 16000.0)."""
        if value in (None, ''):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(',', '').strip())
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _as_int(value):
        try:
            return int(float(str(value).replace(',', '').strip()))
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _week_monday(year, week):
        """Lunes de la semana `week` (1..52): primer lunes del año + (week-1)."""
        from datetime import timedelta
        jan1 = date(year, 1, 1)
        first_monday = jan1 + timedelta(days=(7 - jan1.weekday()) % 7)
        return first_monday + timedelta(weeks=week - 1)

    def _parse_forecast_header(self, ws):
        """Ubica la fila 'SEMANA' y mapea número de semana → columna."""
        header_row = None
        max_r = min(ws.max_row or 0, 40)
        max_c = min(ws.max_column or 0, 120)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                if 'semana' in _norm(ws.cell(r, c).value):
                    header_row = r
                    break
            if header_row:
                break
        if not header_row:
            raise UserError(
                "No se encontró la fila de semanas (debe haber una celda "
                "'SEMANA'). ¿Es la hoja del forecast (F-P-A28-13)?")
        week_cols = {}
        for c in range(1, (ws.max_column or 0) + 1):
            w = self._as_int(ws.cell(header_row, c).value)
            if w and 1 <= w <= 53:
                week_cols[w] = c
        if not week_cols:
            raise UserError(
                "No se reconocieron columnas de semana (números 1–52 en la fila "
                "'SEMANA').")
        return header_row, week_cols

    def _resolve_uom(self, unit_text, product, warnings):
        """Unidad de la columna UNIDAD (validando categoría) o la de venta."""
        if not unit_text:
            return product.uom_id
        uom = self.env['uom.uom'].search([('name', '=', unit_text.strip())], limit=1)
        if not uom:
            uom = self.env['uom.uom'].search([('name', 'ilike', unit_text.strip())], limit=1)
        if not uom:
            warnings.append("Unidad '%s' no reconocida para %s: se usó %s." % (
                unit_text, product.display_name, product.uom_id.name))
            return product.uom_id
        if not product.uom_id._has_common_reference(uom):
            warnings.append("Unidad '%s' de otra categoría para %s: se usó %s." % (
                unit_text, product.display_name, product.uom_id.name))
            return product.uom_id
        return uom

    # --- Importación ---------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        budget = self.budget_id
        if budget.state != 'borrador':
            raise UserError(
                "Solo se importa sobre un presupuesto en borrador.")
        wb = self._load_workbook()
        name = self.sheet_name or (wb.sheetnames[0] if wb.sheetnames else None)
        if not name or name not in wb.sheetnames:
            raise UserError("La hoja '%s' no existe en el archivo." % (name or ''))
        ws = wb[name]
        # Errores ESTRUCTURALES abortan todo (dentro del savepoint). El layout se
        # elige por el tipo del presupuesto (mensual F-P-A28-18 / semanal A28-13).
        with self.env.cr.savepoint():
            if budget.kind == 'pronostico':
                imported, unmatched, warnings = self._import_forecast(ws, budget)
            else:
                imported, unmatched, warnings = self._import_monthly(ws, budget)
            budget.message_post(body=self._result_body(
                imported, unmatched, warnings))
        self.result = self._result_body(imported, unmatched, warnings, html=False)
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sgi.sales.budget.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _import_monthly(self, ws, budget):
        """Layout F-P-A28-18: producto × pares '<mes> m'/'<mes> $'."""
        header_row, prod_col, unit_col, client_col, month_cols = \
            self._parse_header(ws)
        imported, unmatched, warnings, cleared = 0, [], [], set()
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            ref = ws.cell(r, prod_col).value
            if ref is None or str(ref).strip() == '':
                continue
            product = self._match_product(ref)
            if not product:
                unmatched.append(str(ref).strip())
                continue
            unit_text = ws.cell(r, unit_col).value if unit_col else None
            uom = self._resolve_uom(unit_text, product, warnings)
            partner = False
            if client_col:
                ctext = ws.cell(r, client_col).value
                if ctext and str(ctext).strip():
                    partner = self.env['res.partner'].search(
                        [('name', 'ilike', str(ctext).strip())], limit=1)
            if self.conflict_mode == 'replace' and product.id not in cleared:
                budget.line_ids.filtered(
                    lambda l: l.product_id == product).unlink()
                cleared.add(product.id)
            months = {}
            for (month, kind), col in month_cols.items():
                val = self._as_number(ws.cell(r, col).value)
                if val > 0:
                    months.setdefault(month, {})[kind] = val
            for month, kv in months.items():
                if self._upsert_line(
                        budget, product, date(budget.year, month, 1), uom,
                        partner and partner.id, kv.get('qty', 0.0),
                        kv.get('amount', 0.0), warnings):
                    imported += 1
        return imported, unmatched, warnings

    def _import_forecast(self, ws, budget):
        """Layout forecast.xlsx (F-P-A28-13): fila 'SEMANA' con números 1–52,
        producto en col A y código de cliente en col B. Los bloques repetidos del
        mismo producto (oleadas de PO) se SUMAN por semana; las filas PO/TOTAL/
        FECHA (o producto vacío) se ignoran."""
        header_row, week_cols = self._parse_forecast_header(ws)
        unmatched, warnings = [], []
        # Acumula por producto: {product: {'code':.., 'weeks':{w: qty}}}.
        acc = {}
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            ref = ws.cell(r, 1).value
            norm = _norm(ref)
            if not norm or norm == 'po' or 'total' in norm or 'fecha' in norm:
                continue  # filas PO / total / fecha de entrega
            product = self._match_product(ref)
            if not product:
                unmatched.append(str(ref).strip())
                continue
            code = ws.cell(r, 2).value
            code = str(code).strip() if code not in (None, '') else False
            entry = acc.setdefault(product.id, {
                'product': product, 'code': code, 'weeks': {}})
            if code and not entry['code']:
                entry['code'] = code
            for week, col in week_cols.items():
                val = self._as_number(ws.cell(r, col).value)
                if val > 0:
                    entry['weeks'][week] = entry['weeks'].get(week, 0.0) + val
        imported = 0
        cleared = set()
        for entry in acc.values():
            product = entry['product']
            if self.conflict_mode == 'replace' and product.id not in cleared:
                budget.line_ids.filtered(
                    lambda l: l.product_id == product).unlink()
                cleared.add(product.id)
            for week, qty in entry['weeks'].items():
                monday = self._week_monday(budget.year, week)
                if self._upsert_line(
                        budget, product, monday, product.uom_id,
                        budget.partner_id.id, qty, 0.0, warnings,
                        customer_code=entry['code']):
                    imported += 1
        return imported, unmatched, warnings

    def _upsert_line(self, budget, product, line_date, uom, partner_id, qty,
                     amount, warnings, customer_code=None):
        """Crea/suma una línea (producto, fecha, cliente). Errores de datos
        (unidad, esquema mixto, semana fuera de año) se reportan y NO abortan la
        hoja (savepoint por línea)."""
        Line = self.env['sgi.sales.budget.line']
        domain = [('budget_id', '=', budget.id), ('product_id', '=', product.id),
                  ('date', '=', line_date),
                  ('partner_id', '=', partner_id or False)]
        try:
            with self.env.cr.savepoint():
                existing = Line.search(domain, limit=1)
                if existing and self.conflict_mode == 'add':
                    existing.qty_budget += qty
                    if amount:
                        existing.amount_budget = existing.amount_budget + amount
                    return True
                vals = {
                    'budget_id': budget.id, 'product_id': product.id,
                    'date': line_date, 'uom_id': uom.id,
                    'partner_id': partner_id or False, 'qty_budget': qty,
                }
                if customer_code:
                    vals['customer_code'] = customer_code
                if amount:
                    vals['amount_budget'] = amount
                line = Line.create(vals)
                # m sin $: sugiere el importe con el precio de lista.
                if qty and not amount:
                    price, source = line._sgi_suggest_price()
                    if price:
                        line.price_unit_budget = price
                        line.price_source = source
                return True
        except (ValidationError, UserError) as exc:
            warnings.append("%s %02d/%s: %s" % (
                product.display_name, month, budget.year,
                exc.args[0] if exc.args else exc))
            return False

    def _result_body(self, imported, unmatched, warnings, html=True):
        parts = ["%d línea(s) importada(s)." % imported]
        if unmatched:
            parts.append("%d producto(s) sin match: %s" % (
                len(unmatched), ", ".join(unmatched)))
        if warnings:
            parts.append("Avisos:\n- " + "\n- ".join(warnings))
        text = "\n".join(parts)
        if html:
            return text.replace("\n", "<br/>")
        return text
